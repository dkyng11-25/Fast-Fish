#!/usr/bin/env python3
"""
Step 27: Cluster × Role Gap Matrix Generator

This step creates gap analysis matrices showing which product roles 
are missing or underrepresented in each cluster, with Excel formatting.

Builds on Steps 25 (Product Roles) and 26 (Price Bands).
Uses ONLY real data with methodical validation.

Author: Data Pipeline Team
Date: 2025-01-23
Version: 1.0 - Methodical Implementation

 HOW TO RUN (CLI + ENV)
 ----------------------
 Overview
 - Period-aware: this step resolves period-labeled inputs via the manifest where possible.
 - Sales loading uses src.config.get_current_period() unless you override the source via ENV.
 - For fast testing, upstream Steps 25–26 can be run for a single cluster (e.g., Cluster 22), and
   Step 27 will analyze only the clusters present in those inputs. For production, process all clusters.

 Quick Start (fast testing; target 202510A, source 202509A)
   ENV:
     PIPELINE_YYYYMM=202509 PIPELINE_PERIOD=A

   Command:
     PYTHONPATH=. python3 src/step27_gap_matrix_generator.py \
       --target-yyyymm 202510 \
       --target-period A

 Production Run (all clusters)
   Command:
     PYTHONPATH=. python3 src/step27_gap_matrix_generator.py \
       --target-yyyymm 202510 \
       --target-period A

 Period Handling Patterns
 - Source sales period (loader): from PIPELINE_YYYYMM/PIPELINE_PERIOD
 - Output labeling (filenames): from --target-yyyymm/--target-period

 Best Practices & Pitfalls
 - Ensure Step 25 roles and Step 26 price bands exist for the target label (period-aware files or defaults).
 - If you see "No valid period-specific sales files ...", set PIPELINE_YYYYMM/PIPELINE_PERIOD to a prior real period.
 - Avoid synthetic combined files; the loader forbids them.
"""

import pandas as pd
import numpy as np
import os
import json
from datetime import datetime
from typing import Dict, Tuple, Any, List, Optional
import warnings
from tqdm import tqdm
from src.config import get_period_label
from src.pipeline_manifest import get_manifest, register_step_output

# Excel formatting dependencies
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Excel formatting not available (install openpyxl for enhanced features)")

# Suppress pandas warnings
warnings.filterwarnings('ignore')

# ===== CONFIGURATION =====
# Input data sources (period-aware resolvers prefer manifest)
SALES_DATA_FILE = None  # Period-aware via config helpers; do not use combined/synthetic
PRODUCT_ROLES_FILE = "output/product_role_classifications.csv"  # Fallback only
PRICE_BANDS_FILE = "output/price_band_analysis.csv"  # Fallback only
CLUSTER_LABELS_FILE = "output/clustering_results_spu.csv"  # Fallback only

# Output base filenames (period-specific filenames are composed in main)
GAP_MATRIX_EXCEL = "output/gap_matrix.xlsx"
GAP_ANALYSIS_CSV = "output/gap_analysis_detailed.csv"
GAP_SUMMARY_JSON = "output/gap_matrix_summary.json"
GAP_REPORT_FILE = "output/gap_matrix_analysis_report.md"

# Create output directory
os.makedirs("output", exist_ok=True)

# ===== GAP ANALYSIS CONFIGURATION =====
GAP_THRESHOLDS = {
    'critical_gap': 10,  # Gap > 10% is critical (red)
    'moderate_gap': 5,   # Gap 5-10% is moderate (yellow)
    'optimal_range': 5   # Within 5% is optimal (green)
}

# Expected product role distribution (business rules)
EXPECTED_ROLE_DISTRIBUTION = {
    'CORE': {'min': 15, 'max': 35, 'target': 25},      # 15-35%, target 25%
    'SEASONAL': {'min': 20, 'max': 40, 'target': 30},  # 20-40%, target 30%
    'FILLER': {'min': 25, 'max': 45, 'target': 35},   # 25-45%, target 35%
    'CLEARANCE': {'min': 0, 'max': 15, 'target': 10}   # 0-15%, target 10%
}

def log_progress(message: str) -> None:
    """Log progress with timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")

# ===== DATA LOADING AND PREPARATION =====

def _resolve_product_roles_path(period_label: Optional[str]) -> str:
    override = os.getenv("STEP27_PRODUCT_ROLES_FILE")
    if override and os.path.exists(override):
        return override
    try:
        manifest = get_manifest()
        step = (manifest or {}).get("step25", {})
        if period_label:
            key = f"product_role_classifications_{period_label}"
            val = step.get(key)
            path = val.get("path") if isinstance(val, dict) else val
            if path and os.path.exists(path):
                return path
        val = step.get("product_role_classifications")
        path = val.get("path") if isinstance(val, dict) else val
        if path and os.path.exists(path):
            return path
    except Exception:
        pass
    return PRODUCT_ROLES_FILE

def _resolve_price_bands_path(period_label: Optional[str]) -> str:
    override = os.getenv("STEP27_PRICE_BANDS_FILE")
    if override and os.path.exists(override):
        return override
    try:
        manifest = get_manifest()
        step = (manifest or {}).get("step26", {})
        if period_label:
            key = f"price_band_analysis_{period_label}"
            val = step.get(key)
            path = val.get("path") if isinstance(val, dict) else val
            if path and os.path.exists(path):
                return path
        val = step.get("price_band_analysis")
        path = val.get("path") if isinstance(val, dict) else val
        if path and os.path.exists(path):
            return path
    except Exception:
        pass
    return PRICE_BANDS_FILE

def _resolve_cluster_labels_path(period_label: Optional[str]) -> Optional[str]:
    # Prefer period-labeled clustering results if present, then generic
    candidates: List[str] = []
    if period_label:
        candidates += [
            f"output/clustering_results_spu_{period_label}.csv",
            f"output/comprehensive_cluster_labels_{period_label}.csv",
        ]
    # Fallbacks
    candidates += [
        "output/clustering_results_spu.csv",
        "output/comprehensive_cluster_labels.csv",
    ]
    for path in candidates:
        if path and os.path.exists(path):
            return path
    return None

def _resolve_step18_sell_through_path(period_label: Optional[str]) -> Optional[str]:
    """Resolve Step 18 sell-through analysis output via manifest (period-aware)."""
    try:
        man = get_manifest() or {}
        step = man.get("step18", {})
        if period_label:
            k = f"sell_through_analysis_{period_label}"
            v = step.get(k)
            p = v.get("path") if isinstance(v, dict) else v
            if p and os.path.exists(p):
                return p
        v = step.get("sell_through_analysis")
        p = v.get("path") if isinstance(v, dict) else v
        if p and os.path.exists(p):
            return p
    except Exception:
        pass
    return None

def _compute_sell_through_summary(step18_path: str) -> pd.DataFrame:
    """Compute Subcategory-level sell-through summary from Step 18 output (percent)."""
    try:
        df = pd.read_csv(step18_path)
        # Prefer percent column; otherwise convert fraction to percent
        if 'Sell_Through_Rate_Pct' in df.columns:
            st_pct = pd.to_numeric(df['Sell_Through_Rate_Pct'], errors='coerce')
        elif 'Sell_Through_Rate' in df.columns:
            st_pct = pd.to_numeric(df['Sell_Through_Rate'], errors='coerce')
        elif 'Sell_Through_Rate_Frac' in df.columns:
            st_pct = pd.to_numeric(df['Sell_Through_Rate_Frac'], errors='coerce') * 100.0
        else:
            return pd.DataFrame()
        cat = df.get('Category') if 'Category' in df.columns else pd.Series([None] * len(df))
        sub = df.get('Subcategory') if 'Subcategory' in df.columns else pd.Series([None] * len(df))
        s = pd.DataFrame({'Category': cat, 'Subcategory': sub, 'Sell_Through_Pct': st_pct})
        s = s.dropna(subset=['Sell_Through_Pct'])
        if s.empty:
            return pd.DataFrame()
        out = (
            s.groupby(['Category', 'Subcategory'])['Sell_Through_Pct']
             .agg(['count', 'mean', 'median', 'min', 'max'])
             .reset_index()
             .rename(columns={'count': 'Records', 'mean': 'Avg_Sell_Through_Pct', 'median': 'Median_Sell_Through_Pct',
                              'min': 'Min_Sell_Through_Pct', 'max': 'Max_Sell_Through_Pct'})
        )
        # Round for presentation
        for c in ['Avg_Sell_Through_Pct', 'Median_Sell_Through_Pct', 'Min_Sell_Through_Pct', 'Max_Sell_Through_Pct']:
            out[c] = out[c].round(1)
        return out
    except Exception:
        return pd.DataFrame()

def load_and_prepare_data(period_label: Optional[str] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load and prepare all required data sources"""
    log_progress("🔍 Loading data for gap matrix analysis...")
    
    # Resolve period-aware inputs (manifest-backed where possible)
    roles_path = _resolve_product_roles_path(period_label)
    price_bands_path = _resolve_price_bands_path(period_label)
    cluster_labels_path = _resolve_cluster_labels_path(period_label)
    if not roles_path or not os.path.exists(roles_path):
        raise FileNotFoundError(f"Product roles not found: {roles_path}")
    if not price_bands_path or not os.path.exists(price_bands_path):
        raise FileNotFoundError(f"Price bands not found: {price_bands_path}")
    if not cluster_labels_path or not os.path.exists(cluster_labels_path):
        raise FileNotFoundError(f"Cluster labels not found: {cluster_labels_path}")
    
    # Load period-specific SPU-level sales (NO synthetic combined files)
    from src.config import get_current_period, load_sales_df_with_fashion_basic
    yyyymm, period = get_current_period()
    src_path, sales_df = load_sales_df_with_fashion_basic(yyyymm, period, require_spu_level=True,
        required_cols=['str_code','spu_code','fashion_sal_amt','basic_sal_amt','cate_name','sub_cate_name'])
    log_progress(f"   ✓ Loaded period-specific sales: {src_path} ({len(sales_df):,} records)")
    
    # Validate sales data columns
    required_sales_cols = ['str_code', 'spu_code', 'fashion_sal_amt', 'basic_sal_amt', 'cate_name', 'sub_cate_name']
    missing_sales_cols = [col for col in required_sales_cols if col not in sales_df.columns]
    if missing_sales_cols:
        raise ValueError(f"Missing required columns in sales data: {missing_sales_cols}")
    
    # Load product role classifications
    product_roles_df = pd.read_csv(roles_path)
    log_progress(f"   ✓ Loaded product roles: {len(product_roles_df):,} products")
    
    # Validate product roles columns
    required_roles_cols = ['spu_code', 'product_role', 'category', 'subcategory']
    missing_roles_cols = [col for col in required_roles_cols if col not in product_roles_df.columns]
    if missing_roles_cols:
        raise ValueError(f"Missing required columns in product roles: {missing_roles_cols}")
    
    # Load price band analysis
    price_bands_df = pd.read_csv(price_bands_path)
    log_progress(f"   ✓ Loaded price bands: {len(price_bands_df):,} products")
    
    # Validate price bands columns
    required_price_cols = ['spu_code', 'price_band', 'avg_unit_price']
    missing_price_cols = [col for col in required_price_cols if col not in price_bands_df.columns]
    if missing_price_cols:
        raise ValueError(f"Missing required columns in price bands: {missing_price_cols}")
    
    # Load and process cluster information
    cluster_df = pd.read_csv(cluster_labels_path)
    log_progress(f"   ✓ Loaded cluster data: {len(cluster_df):,} clusters")
    
    # Normalize cluster column naming to a standard 'cluster_id'
    if 'cluster_id' not in cluster_df.columns:
        if 'Cluster' in cluster_df.columns:
            cluster_df = cluster_df.rename(columns={'Cluster': 'cluster_id'})
            log_progress("   📝 Normalized cluster column: 'Cluster' → 'cluster_id'")
        elif 'cluster' in cluster_df.columns:
            cluster_df = cluster_df.rename(columns={'cluster': 'cluster_id'})
            log_progress("   📝 Normalized cluster column: 'cluster' → 'cluster_id'")
        elif 'cluster_label' in cluster_df.columns:
            cluster_df = cluster_df.rename(columns={'cluster_label': 'cluster_id'})
            log_progress("   📝 Normalized cluster column: 'cluster_label' → 'cluster_id'")
    
    # Convert cluster-level data to store-level mapping if needed
    if 'str_code' not in cluster_df.columns and 'store_codes' in cluster_df.columns:
        log_progress(f"   📝 Converting cluster-level to store-level mapping...")
        
        store_cluster_mapping = []
        for _, row in cluster_df.iterrows():
            cluster_id = row['cluster_id']
            store_codes_str = str(row['store_codes'])
            
            # Parse store codes
            if '...' in store_codes_str:
                visible_stores = store_codes_str.replace('...', '').split(',')
                visible_stores = [store.strip() for store in visible_stores if store.strip()]
            else:
                visible_stores = store_codes_str.split(',')
                visible_stores = [store.strip() for store in visible_stores if store.strip()]
            
            for store_code in visible_stores:
                if store_code:
                    store_cluster_mapping.append({
                        'str_code': store_code,
                        'cluster_id': cluster_id
                    })
        
        store_cluster_df = pd.DataFrame(store_cluster_mapping)
        store_cluster_df['str_code'] = store_cluster_df['str_code'].astype(str)
        log_progress(f"   ✓ Created store-level mapping: {len(store_cluster_df):,} stores")
    else:
        store_cluster_df = cluster_df[['str_code', 'cluster_id']].copy()
        store_cluster_df['str_code'] = store_cluster_df['str_code'].astype(str)
    
    return sales_df, product_roles_df, price_bands_df, store_cluster_df

# ===== GAP ANALYSIS FUNCTIONS =====

def analyze_cluster_role_distribution(sales_df: pd.DataFrame, product_roles_df: pd.DataFrame, 
                                     store_cluster_df: pd.DataFrame) -> pd.DataFrame:
    """Analyze current product role distribution by cluster"""
    log_progress("📊 Analyzing current cluster × role distribution...")
    
    # Merge sales data with cluster and role information
    sales_with_clusters = sales_df.merge(store_cluster_df, on='str_code', how='inner')
    sales_with_roles = sales_with_clusters.merge(
        product_roles_df[['spu_code', 'product_role', 'category', 'subcategory']], 
        on='spu_code', 
        how='inner'
    )
    
    log_progress(f"   ✓ Merged data: {len(sales_with_roles):,} records with cluster and role info")
    
    # Calculate current distribution
    cluster_role_stats = []
    
    for cluster_id in tqdm(sorted(sales_with_roles['cluster_id'].unique()), desc="Analyzing clusters"):
        cluster_data = sales_with_roles[sales_with_roles['cluster_id'] == cluster_id]
        total_products = len(cluster_data)
        
        if total_products == 0:
            continue
        
        cluster_stats = {
            'cluster_id': cluster_id,
            'total_products': total_products,
            'total_stores': cluster_data['str_code'].nunique()
        }
        
        # Calculate role distribution
        role_counts = cluster_data['product_role'].value_counts()
        for role in ['CORE', 'SEASONAL', 'FILLER', 'CLEARANCE']:
            count = role_counts.get(role, 0)
            percentage = (count / total_products) * 100 if total_products > 0 else 0
            
            cluster_stats[f'{role.lower()}_count'] = count
            cluster_stats[f'{role.lower()}_percentage'] = percentage
            
            # Calculate gap vs expected
            expected = EXPECTED_ROLE_DISTRIBUTION[role]['target']
            gap = expected - percentage
            gap_severity = classify_gap_severity(gap)
            
            cluster_stats[f'{role.lower()}_gap'] = gap
            cluster_stats[f'{role.lower()}_gap_severity'] = gap_severity
        
        # Calculate category diversity
        cluster_stats['category_count'] = cluster_data['category'].nunique()
        cluster_stats['subcategory_count'] = cluster_data['subcategory'].nunique()
        
        cluster_role_stats.append(cluster_stats)
    
    distribution_df = pd.DataFrame(cluster_role_stats)
    log_progress(f"   ✓ Analyzed {len(distribution_df):,} clusters")
    
    return distribution_df

def classify_gap_severity(gap: float) -> str:
    """Classify gap severity based on thresholds"""
    abs_gap = abs(gap)
    
    if abs_gap >= GAP_THRESHOLDS['critical_gap']:
        return 'CRITICAL'
    elif abs_gap >= GAP_THRESHOLDS['moderate_gap']:
        return 'MODERATE'
    else:
        return 'OPTIMAL'

def create_gap_matrix(distribution_df: pd.DataFrame) -> pd.DataFrame:
    """Create the main gap matrix for visualization"""
    log_progress("🏗️ Creating gap matrix...")
    
    # Create matrix with clusters as rows and roles as columns
    matrix_data = []
    
    for _, cluster in distribution_df.iterrows():
        row = {
            'Cluster_ID': f"Cluster_{cluster['cluster_id']}",
            'Total_Products': cluster['total_products'],
            'Total_Stores': cluster['total_stores']
        }
        
        # Add role percentages and gaps
        for role in ['CORE', 'SEASONAL', 'FILLER', 'CLEARANCE']:
            role_lower = role.lower()
            percentage = cluster[f'{role_lower}_percentage']
            gap = cluster[f'{role_lower}_gap']
            severity = cluster[f'{role_lower}_gap_severity']
            
            row[f'{role}_Current_%'] = percentage
            row[f'{role}_Gap'] = gap
            row[f'{role}_Status'] = severity
        
        matrix_data.append(row)
    
    matrix_df = pd.DataFrame(matrix_data)
    log_progress(f"   ✓ Created gap matrix: {matrix_df.shape}")
    
    return matrix_df

# ===== EXCEL FORMATTING FUNCTIONS =====

def create_formatted_excel(matrix_df: pd.DataFrame, distribution_df: pd.DataFrame, excel_out_path: str, sell_through_summary: Optional[pd.DataFrame] = None) -> None:
    """Create Excel file with conditional formatting"""
    log_progress("📋 Creating formatted Excel file...")
    
    if not EXCEL_AVAILABLE:
        log_progress("   ⚠️ Saving as CSV (Excel formatting unavailable)")
        matrix_df.to_csv(excel_out_path.replace('.xlsx', '.csv'), index=False)
        return
    
    # Create workbook and worksheets
    wb = openpyxl.Workbook()
    
    # Sheet 1: Gap Matrix
    ws_matrix = wb.active
    ws_matrix.title = "Gap Matrix"
    
    # Write gap matrix data
    for row in dataframe_to_rows(matrix_df, index=False, header=True):
        ws_matrix.append(row)
    
    # Apply formatting to gap matrix
    apply_gap_matrix_formatting(ws_matrix, matrix_df)
    
    # Sheet 2: Detailed Analysis
    ws_detail = wb.create_sheet("Detailed Analysis")
    for row in dataframe_to_rows(distribution_df, index=False, header=True):
        ws_detail.append(row)
    
    # Sheet 3: Summary & Recommendations
    ws_summary = wb.create_sheet("Summary")
    create_summary_sheet(ws_summary, distribution_df)
    
    # Optional Sheet 3: Sell-Through Summary (from Step 18)
    if sell_through_summary is not None and not sell_through_summary.empty:
        ws_st = wb.create_sheet("Sell-Through Summary")
        for row in dataframe_to_rows(sell_through_summary, index=False, header=True):
            ws_st.append(row)
        # Simple formatting
        for cell in ws_st[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save workbook
    wb.save(excel_out_path)
    log_progress(f"   ✅ Saved formatted Excel: {excel_out_path}")

def apply_gap_matrix_formatting(ws, matrix_df):
    """Apply conditional formatting to gap matrix"""
    if not EXCEL_AVAILABLE:
        return
    
    # Define colors
    critical_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red
    moderate_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Light yellow
    optimal_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")   # Light green
    
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")    # Light blue
    header_font = Font(bold=True)
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Apply conditional formatting based on gap severity
    for row_idx, row in enumerate(matrix_df.itertuples(), start=2):
        for role in ['CORE', 'SEASONAL', 'FILLER', 'CLEARANCE']:
            status_col = None
            gap_col = None
            
            # Find column indices
            for col_idx, header in enumerate(matrix_df.columns, start=1):
                if header == f'{role}_Status':
                    status_col = col_idx
                elif header == f'{role}_Gap':
                    gap_col = col_idx
            
            if status_col and gap_col:
                status_value = getattr(row, f'{role}_Status')
                gap_cell = ws.cell(row=row_idx, column=gap_col)
                
                # Apply color based on status
                if status_value == 'CRITICAL':
                    gap_cell.fill = critical_fill
                elif status_value == 'MODERATE':
                    gap_cell.fill = moderate_fill
                else:
                    gap_cell.fill = optimal_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_summary_sheet(ws, distribution_df):
    """Create summary and recommendations sheet"""
    if not EXCEL_AVAILABLE:
        return
    
    # Add title
    ws.append(['Cluster × Role Gap Analysis Summary'])
    ws.append([])
    ws.append([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    ws.append([])
    
    # Critical gaps summary
    ws.append(['CRITICAL GAPS (>10%):'])
    ws.append(['Cluster', 'Role', 'Gap %', 'Recommendation'])
    
    for _, cluster in distribution_df.iterrows():
        cluster_id = f"Cluster_{cluster['cluster_id']}"
        for role in ['CORE', 'SEASONAL', 'FILLER', 'CLEARANCE']:
            role_lower = role.lower()
            gap = cluster[f'{role_lower}_gap']
            severity = cluster[f'{role_lower}_gap_severity']
            
            if severity == 'CRITICAL':
                if gap > 0:
                    recommendation = f"Need to add {abs(gap):.1f}% more {role} products"
                else:
                    recommendation = f"Consider reducing {abs(gap):.1f}% of {role} products"
                
                ws.append([cluster_id, role, f"{gap:+.1f}%", recommendation])
    
    # Format summary sheet
    ws['A1'].font = Font(bold=True, size=14)
    ws['A5'].font = Font(bold=True)

# ===== REPORTING FUNCTIONS =====

def create_gap_analysis_summary(distribution_df: pd.DataFrame, matrix_df: pd.DataFrame) -> Dict[str, Any]:
    """Create comprehensive gap analysis summary"""
    
    # Count gaps by severity
    total_gaps = 0
    critical_gaps = 0
    moderate_gaps = 0
    
    for role in ['CORE', 'SEASONAL', 'FILLER', 'CLEARANCE']:
        role_lower = role.lower()
        severity_col = f'{role_lower}_gap_severity'
        if severity_col in distribution_df.columns:
            critical_gaps += len(distribution_df[distribution_df[severity_col] == 'CRITICAL'])
            moderate_gaps += len(distribution_df[distribution_df[severity_col] == 'MODERATE'])
            total_gaps += len(distribution_df[distribution_df[severity_col].isin(['CRITICAL', 'MODERATE'])])
    
    summary = {
        'analysis_metadata': {
            'total_clusters': len(distribution_df),
            'total_products_analyzed': distribution_df['total_products'].sum(),
            'analysis_timestamp': datetime.now().isoformat(),
            'gap_analysis_method': 'cluster_role_distribution_vs_expected'
        },
        
        'gap_severity_counts': {
            'critical_gaps': critical_gaps,
            'moderate_gaps': moderate_gaps,
            'total_significant_gaps': total_gaps,
            'optimal_clusters': len(distribution_df) * 4 - total_gaps  # 4 roles per cluster
        },
        
        'cluster_summary': {},
        
        'role_gap_analysis': {},
        
        'recommendations': {
            'high_priority': [],
            'medium_priority': [],
            'optimization_opportunities': []
        }
    }
    
    # Add cluster-specific summaries
    for _, cluster in distribution_df.iterrows():
        cluster_id = f"Cluster_{cluster['cluster_id']}"
        cluster_gaps = []
        
        for role in ['CORE', 'SEASONAL', 'FILLER', 'CLEARANCE']:
            role_lower = role.lower()
            gap = cluster[f'{role_lower}_gap']
            severity = cluster[f'{role_lower}_gap_severity']
            
            if severity in ['CRITICAL', 'MODERATE']:
                cluster_gaps.append({
                    'role': role,
                    'gap': gap,
                    'severity': severity
                })
        
        summary['cluster_summary'][cluster_id] = {
            'total_products': cluster['total_products'],
            'total_stores': cluster['total_stores'],
            'significant_gaps': len(cluster_gaps),
            'gaps_detail': cluster_gaps
        }
    
    # Add role-specific gap analysis
    for role in ['CORE', 'SEASONAL', 'FILLER', 'CLEARANCE']:
        role_lower = role.lower()
        gap_col = f'{role_lower}_gap'
        
        if gap_col in distribution_df.columns:
            role_gaps = distribution_df[gap_col]
            summary['role_gap_analysis'][role] = {
                'avg_gap': float(role_gaps.mean()),
                'max_gap': float(role_gaps.max()),
                'min_gap': float(role_gaps.min()),
                'clusters_with_critical_gaps': len(distribution_df[distribution_df[f'{role_lower}_gap_severity'] == 'CRITICAL'])
            }
    
    return summary

def create_detailed_report(distribution_df: pd.DataFrame, summary: Dict[str, Any]) -> None:
    """Create detailed gap analysis report"""
    
    report_content = f"""# Cluster × Role Gap Matrix Analysis Report

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  
**Clusters Analyzed:** {summary['analysis_metadata']['total_clusters']}  
**Products Analyzed:** {summary['analysis_metadata']['total_products_analyzed']}

## Executive Summary

This report identifies gaps in product role distribution across clusters, highlighting 
where clusters lack appropriate balance of CORE, SEASONAL, FILLER, and CLEARANCE products.

## Gap Severity Overview

- **Critical Gaps (>10%):** {summary['gap_severity_counts']['critical_gaps']} instances
- **Moderate Gaps (5-10%):** {summary['gap_severity_counts']['moderate_gaps']} instances  
- **Total Significant Gaps:** {summary['gap_severity_counts']['total_significant_gaps']} instances
- **Optimal Allocations:** {summary['gap_severity_counts']['optimal_clusters']} instances

## Critical Gaps by Cluster

"""
    
    # Add critical gap details
    for cluster_id, cluster_data in summary['cluster_summary'].items():
        critical_gaps = [gap for gap in cluster_data['gaps_detail'] if gap['severity'] == 'CRITICAL']
        
        if critical_gaps:
            report_content += f"""### {cluster_id}
- **Products:** {cluster_data['total_products']} across {cluster_data['total_stores']} stores
- **Critical Gaps:** {len(critical_gaps)}

"""
            for gap in critical_gaps:
                gap_direction = "shortage" if gap['gap'] > 0 else "excess"
                report_content += f"""  - **{gap['role']}:** {abs(gap['gap']):.1f}% {gap_direction}
"""
            report_content += "\n"
    
    report_content += f"""
## Role-Specific Gap Analysis

"""
    
    for role, role_data in summary['role_gap_analysis'].items():
        expected = EXPECTED_ROLE_DISTRIBUTION[role]['target']
        report_content += f"""### {role} Products (Target: {expected}%)
- **Average Gap:** {role_data['avg_gap']:+.1f}%
- **Largest Gap:** {role_data['max_gap']:+.1f}%
- **Clusters with Critical Gaps:** {role_data['clusters_with_critical_gaps']}

"""
    
    report_content += f"""
## Expected vs Actual Distribution

| Role | Target % | Min % | Max % | Business Rule |
|------|----------|-------|-------|---------------|
"""
    
    for role, config in EXPECTED_ROLE_DISTRIBUTION.items():
        report_content += f"""| {role} | {config['target']}% | {config['min']}% | {config['max']}% | {get_role_description(role)} |
"""
    
    report_content += """
## Methodology

### Gap Calculation
- **Current Distribution:** Actual percentage of each role in each cluster
- **Expected Distribution:** Business-defined target percentages
- **Gap:** Expected % - Current %
- **Severity:** Critical (>10%), Moderate (5-10%), Optimal (<5%)

### Data Sources
- Cluster assignments from comprehensive cluster labeling
- Product role classifications from Step 25
- Sales transaction data for product-store mapping

---
*Generated by Gap Matrix Analyzer v1.0*
"""
    
    with open(GAP_REPORT_FILE, 'w', encoding='utf-8') as f:
        f.write(report_content)
    
    log_progress(f"   ✓ Saved detailed report: {GAP_REPORT_FILE}")

def get_role_description(role: str) -> str:
    """Get business description for each role"""
    descriptions = {
        'CORE': 'High-volume, consistent performers',
        'SEASONAL': 'Fashion-driven, time-sensitive',
        'FILLER': 'Gap-filling, moderate volume',
        'CLEARANCE': 'End-of-life, discounted items'
    }
    return descriptions.get(role, 'Unknown role')

def main() -> None:
    """Main function for gap matrix generation"""
    global GAP_REPORT_FILE  # Fix: Move global declaration to the beginning
    start_time = datetime.now()
    log_progress("🚀 Starting Gap Matrix Generation (Step 27)...")
    
    try:
        # Period-aware CLI
        import argparse
        parser = argparse.ArgumentParser(description="Step 27: Gap Matrix Generator (period-aware)")
        parser.add_argument("--target-yyyymm", required=False)
        parser.add_argument("--target-period", required=False, choices=["A","B"])
        args, _ = parser.parse_known_args()
        period_label = None
        if getattr(args, 'target_yyyymm', None) and getattr(args, 'target_period', None):
            period_label = get_period_label(args.target_yyyymm, args.target_period)

        # Step 1: Load and prepare data
        sales_df, product_roles_df, price_bands_df, store_cluster_df = load_and_prepare_data(period_label=period_label)
        
        # Step 2: Analyze cluster-role distribution
        distribution_df = analyze_cluster_role_distribution(sales_df, product_roles_df, store_cluster_df)
        
        # Step 3: Create gap matrix
        matrix_df = create_gap_matrix(distribution_df)
        
        # Step 4: Save results (DUAL OUTPUT PATTERN - both timestamped and generic)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Always create timestamped version (for backup/inspection)
        if period_label:
            timestamped_gap_analysis_out = f"output/gap_analysis_detailed_{period_label}_{ts}.csv"
            distribution_df.to_csv(timestamped_gap_analysis_out, index=False)
            log_progress(f"✅ Saved timestamped gap analysis: {timestamped_gap_analysis_out}")
        
        # Create symlink for generic version (for pipeline flow)
        generic_gap_analysis_out = GAP_ANALYSIS_CSV
        if os.path.exists(generic_gap_analysis_out) or os.path.islink(generic_gap_analysis_out):
            os.remove(generic_gap_analysis_out)
        os.symlink(os.path.basename(timestamped_gap_analysis_out), generic_gap_analysis_out)
        log_progress(f"✅ Created symlink: {generic_gap_analysis_out} -> {timestamped_gap_analysis_out}")
        
        # Use timestamped version for manifest registration if available, otherwise generic
        gap_analysis_out = timestamped_gap_analysis_out if period_label else generic_gap_analysis_out
        
        # Step 5: Optionally include Step 18 sell-through summary
        st_path = _resolve_step18_sell_through_path(period_label)
        sell_through_summary = _compute_sell_through_summary(st_path) if st_path else pd.DataFrame()
        # Save summary to CSV if present
        st_summary_out = None
        if sell_through_summary is not None and not sell_through_summary.empty:
            st_summary_out = (
                f"output/sell_through_summary_{period_label}_{ts}.csv" if period_label else
                "output/sell_through_summary.csv"
            )
            sell_through_summary.to_csv(st_summary_out, index=False)
            log_progress(f"✅ Saved sell-through summary: {st_summary_out}")

        # Create formatted Excel
        excel_out = (
            f"output/gap_matrix_{period_label}_{ts}.xlsx" if period_label else
            GAP_MATRIX_EXCEL
        )
        create_formatted_excel(matrix_df, distribution_df, excel_out, sell_through_summary=sell_through_summary)
        
        # Step 6: Create summary and report
        summary = create_gap_analysis_summary(distribution_df, matrix_df)
        
        # Convert numpy types to native Python types for JSON serialization
        import json
        def convert_numpy_types(obj):
            if isinstance(obj, np.integer):
                return int(obj)
            elif isinstance(obj, np.floating):
                return float(obj)
            elif isinstance(obj, np.ndarray):
                return obj.tolist()
            elif isinstance(obj, dict):
                return {key: convert_numpy_types(value) for key, value in obj.items()}
            elif isinstance(obj, list):
                return [convert_numpy_types(item) for item in obj]
            return obj
        
        summary_clean = convert_numpy_types(summary)
        
        # Save summary (DUAL OUTPUT PATTERN)
        # Always create timestamped version (for backup/inspection)
        if period_label:
            timestamped_summary_out = f"output/gap_matrix_summary_{period_label}_{ts}.json"
            with open(timestamped_summary_out, 'w') as f:
                json.dump(summary_clean, f, indent=2)
            log_progress(f"✅ Saved timestamped gap summary: {timestamped_summary_out}")
        
        # Always create generic version (for pipeline flow)
        generic_summary_out = GAP_SUMMARY_JSON
        with open(generic_summary_out, 'w') as f:
            json.dump(summary_clean, f, indent=2)
        log_progress(f"✅ Saved generic gap summary: {generic_summary_out}")
        
        # Use timestamped version for manifest registration if available, otherwise generic
        summary_out = timestamped_summary_out if period_label else generic_summary_out
        
        # Create analysis report (DUAL OUTPUT PATTERN)
        # Always create timestamped version (for backup/inspection)
        if period_label:
            timestamped_report_out = f"output/gap_matrix_analysis_report_{period_label}_{ts}.md"
            _old_report = GAP_REPORT_FILE
            GAP_REPORT_FILE = timestamped_report_out
            create_detailed_report(distribution_df, summary)
            GAP_REPORT_FILE = _old_report
            log_progress(f"✅ Saved timestamped analysis report: {timestamped_report_out}")
        
        # Always create generic version (for pipeline flow)
        generic_report_out = GAP_REPORT_FILE
        _old_report = GAP_REPORT_FILE
        GAP_REPORT_FILE = generic_report_out
        create_detailed_report(distribution_df, summary)
        GAP_REPORT_FILE = _old_report
        log_progress(f"✅ Saved generic analysis report: {generic_report_out}")
        
        # Use timestamped version for manifest registration if available, otherwise generic
        report_out = timestamped_report_out if period_label else generic_report_out

        # Register outputs in manifest
        try:
            meta = {
                "records": int(len(distribution_df)),
            }
            register_step_output("step27", "gap_matrix_excel", excel_out, meta)
            register_step_output("step27", "gap_analysis_detailed", gap_analysis_out, meta)
            register_step_output("step27", "gap_matrix_summary", summary_out, {})
            register_step_output("step27", "gap_matrix_analysis_report", report_out, {})
            if st_summary_out:
                register_step_output("step27", "sell_through_summary", st_summary_out, {})
            if period_label:
                register_step_output("step27", f"gap_matrix_excel_{period_label}", excel_out, meta)
                register_step_output("step27", f"gap_analysis_detailed_{period_label}", gap_analysis_out, meta)
                register_step_output("step27", f"gap_matrix_summary_{period_label}", summary_out, {})
                register_step_output("step27", f"gap_matrix_analysis_report_{period_label}", report_out, {})
                if st_summary_out:
                    register_step_output("step27", f"sell_through_summary_{period_label}", st_summary_out, {})
        except Exception:
            pass
        
        # Print final results
        execution_time = (datetime.now() - start_time).total_seconds()
        log_progress(f"\n🎯 GAP MATRIX ANALYSIS RESULTS:")
        log_progress(f"   📊 Clusters Analyzed: {len(distribution_df):,}")
        log_progress(f"   🔍 Gap Analysis:")
        log_progress(f"     • Critical Gaps: {summary['gap_severity_counts']['critical_gaps']}")
        log_progress(f"     • Moderate Gaps: {summary['gap_severity_counts']['moderate_gaps']}")
        log_progress(f"     • Total Products: {summary['analysis_metadata']['total_products_analyzed']:,}")
        
        log_progress(f"   ⚡ Execution Time: {execution_time:.1f} seconds")
        log_progress(f"\n📁 Generated Files:")
        log_progress(f"   • {excel_out}")
        log_progress(f"   • {gap_analysis_out}")
        log_progress(f"   • {summary_out}")
        log_progress(f"   • {report_out}")
        
        log_progress(f"\n✅ Gap Matrix Generation completed successfully")
        
    except Exception as e:
        log_progress(f"❌ Error in gap matrix generation: {str(e)}")
        raise

if __name__ == "__main__":
    main() 