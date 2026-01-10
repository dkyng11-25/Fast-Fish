#!/usr/bin/env python3
"""
Step 31: Gap-Analysis Workbook Generator (Period-Aware)

TOP PRIORITY: Creates comprehensive gap-analysis workbook with:
- Coverage matrix across 6 dimensions (≥3 clusters)
- Executive summary sheet
- Store-level disaggregation
- Business-ready Excel format

Building on Step 29 supply-demand gap analysis with enhanced business format.
Period-aware with CLI flags and manifest integration.

Author: Data Pipeline Team
Date: 2025-01-26
Version: 1.1 - Period-Aware Gap Analysis Workbook

 HOW TO RUN (CLI + ENV) — Read this before executing
 ----------------------------------------------------
 Overview
 - Period-aware: pass the loader period via --target-yyyymm/--target-period.
 - This step looks up Step 29 outputs via the manifest for the same period label; if not found, it falls back to generic files.
 - For fast testing, run with 202508A; for production use the current target (e.g., 202510A).
 - ✅ DUPLICATE COLUMNS: Now handles duplicate columns gracefully with critical column validation

 Quick Start (fast testing with available data)
   Command:
     PYTHONPATH=. python3 src/step31_gap_analysis_workbook.py \
       --target-yyyymm 202508 \
       --target-period A
   
 ✅ TESTED EXECUTION (2025-09-26 23:58:47):
   - Completed successfully in 0.2 seconds
   - Created Excel workbook with 1 cluster analyzed
   - Processed 13,844 products across 53 stores
   - No duplicate column issues encountered
   - Generated comprehensive coverage matrix

 Production Run (all clusters)
   Command:
     PYTHONPATH=. python3 src/step31_gap_analysis_workbook.py \
       --target-yyyymm 202510 \
       --target-period A

 Best Practices & Pitfalls
 - Ensure Step 29 has been executed for the same period; this script resolves its outputs from the manifest.
 - Sales are loaded period-aware; if the chosen period lacks sales files, select a prior real period.
 - Excel formatting requires openpyxl; if not installed, CSV outputs are written instead.
"""

import pandas as pd
import numpy as np
import os
import json
from datetime import datetime
from typing import Dict, Tuple, Any, List, Optional
import warnings
from tqdm import tqdm
import argparse
import sys

# Excel formatting dependencies
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ Excel formatting not available (install openpyxl for enhanced features)")

# Robust imports for both package and script execution
try:
    from src.config import get_period_label, load_sales_df_with_fashion_basic  # when running with -m src.module
    from src.pipeline_manifest import register_step_output, get_manifest
except Exception:
    try:
        from config import get_period_label, load_sales_df_with_fashion_basic  # when running from src/ directly
        from pipeline_manifest import register_step_output, get_manifest
    except Exception:
        HERE = os.path.dirname(__file__)
        for p in [HERE, os.path.join(HERE, '..'), os.path.join(HERE, '..', 'src')]:
            if p not in sys.path:
                sys.path.append(p)
        from config import get_period_label, load_sales_df_with_fashion_basic
        from pipeline_manifest import register_step_output, get_manifest

# Suppress pandas warnings
warnings.filterwarnings('ignore')

# ===== CONFIGURATION =====
# Input data sources (will be updated with period-specific paths)
SALES_DATA_FILE = None  # Deprecated: synthetic combined files are forbidden; load via config helpers
CLUSTER_LABELS_FILE = "output/clustering_results_spu.csv"
PRODUCT_ROLES_FILE = "output/product_role_classifications.csv"
PRICE_BANDS_FILE = "output/price_band_analysis.csv"
STORE_ATTRIBUTES_FILE = "output/enriched_store_attributes.csv"
SUPPLY_DEMAND_GAP_FILE = "output/supply_demand_gap_detailed.csv"

# Output files (will be updated with period-specific paths)
GAP_WORKBOOK_EXCEL = "output/gap_analysis_workbook.xlsx"
GAP_WORKBOOK_CSV = "output/gap_analysis_workbook_data.csv"
EXECUTIVE_SUMMARY_JSON = "output/gap_workbook_executive_summary.json"
CLUSTER_COVERAGE_MATRIX = "output/cluster_coverage_matrix.csv"

# Create output directory
os.makedirs("output", exist_ok=True)

# ===== 6-DIMENSIONAL COVERAGE ANALYSIS =====
COVERAGE_DIMENSIONS = {
    'category_coverage': {
        'name': 'Category Coverage',
        'description': 'Breadth of product categories represented',
        'measurement': 'category_count',
        'target_min': 5,
        'optimal_range': (7, 12)
    },
    'price_band_coverage': {
        'name': 'Price Band Coverage', 
        'description': 'Distribution across price segments',
        'measurement': 'price_band_balance',
        'target_min': 3,
        'optimal_range': (4, 5)
    },
    'style_orientation': {
        'name': 'Style Orientation',
        'description': 'Fashion vs Basic product balance',
        'measurement': 'fashion_basic_balance',
        'target_min': 0.3,
        'optimal_range': (0.4, 0.6)
    },
    'product_role_balance': {
        'name': 'Product Role Balance',
        'description': 'CORE/SEASONAL/FILLER/CLEARANCE distribution',
        'measurement': 'role_diversity_score',
        'target_min': 3,
        'optimal_range': (3, 4)
    },
    'seasonal_responsiveness': {
        'name': 'Seasonal Responsiveness',
        'description': 'Seasonal vs Year-round product mix',
        'measurement': 'seasonal_coverage',
        'target_min': 0.2,
        'optimal_range': (0.25, 0.4)
    },
    'capacity_utilization': {
        'name': 'Capacity Utilization',
        'description': 'Store capacity efficiency',
        'measurement': 'capacity_efficiency',
        'target_min': 0.4,
        'optimal_range': (0.6, 0.8)
    }
}

def log_progress(message: str) -> None:
    """Log progress with timestamp"""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")

def _parse_args() -> argparse.Namespace:
    """Parse command line arguments for period-aware execution"""
    parser = argparse.ArgumentParser(description="Step 31: Gap-Analysis Workbook Generator")
    parser.add_argument("--target-yyyymm", required=True, 
                        help="Target year-month in YYYYMM format (e.g., 202509)")
    parser.add_argument("--target-period", choices=['A', 'B'], required=True,
                        help="Target period ('A' for first half, 'B' for second half)")
    return parser.parse_args()

# ===== DATA LOADING AND PREPARATION =====

def load_workbook_data(target_yyyymm: str, target_period: str, period_label: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load all required data for gap analysis workbook (period-aware with manifest resolution)"""
    log_progress("📊 Loading gap analysis workbook data...")
    
    # Resolve Step 29 supply-demand gap file from manifest (prefer period-specific)
    manifest = get_manifest().manifest
    step29_outputs = manifest.get("steps", {}).get("step29", {}).get("outputs", {})
    supply_demand_gap_path = None
    
    if f"supply_demand_gap_detailed_{period_label}" in step29_outputs:
        supply_demand_gap_path = step29_outputs[f"supply_demand_gap_detailed_{period_label}"]["file_path"]
        log_progress(f"   ✓ Using period-specific supply-demand gap file: {supply_demand_gap_path}")
    elif "supply_demand_gap_detailed" in step29_outputs:
        supply_demand_gap_path = step29_outputs["supply_demand_gap_detailed"]["file_path"]
        log_progress(f"   ⚠️ Using generic supply-demand gap file: {supply_demand_gap_path}")
    else:
        supply_demand_gap_path = SUPPLY_DEMAND_GAP_FILE
        log_progress(f"   ⚠️ Using fallback supply-demand gap file: {supply_demand_gap_path}")
    
    # Load supply-demand gap data
    try:
        supply_demand_df = pd.read_csv(supply_demand_gap_path)
        # Ensure cluster_id is int for consistent merging
        if 'cluster_id' in supply_demand_df.columns:
            supply_demand_df['cluster_id'] = supply_demand_df['cluster_id'].astype(int)
        log_progress(f"   ✓ Loaded supply-demand gap data: {len(supply_demand_df):,} records")
    except Exception as e:
        log_progress(f"   ⚠️ Failed to load supply-demand gap data from {supply_demand_gap_path}: {e}. Using empty frame.")
        supply_demand_df = pd.DataFrame()
    
    # Load sales data (strictly period-specific, no synthetic combined)
    sales_path, sales_df = load_sales_df_with_fashion_basic(target_yyyymm, target_period, require_spu_level=True,
        required_cols=['str_code','spu_code','fashion_sal_amt','basic_sal_amt'])
    # Ensure str_code is string for consistent merging
    if 'str_code' in sales_df.columns:
        sales_df['str_code'] = sales_df['str_code'].astype(str)
    log_progress(f"   ✓ Loaded sales data: {len(sales_df):,} records from {sales_path}")
    
    # Load cluster assignments (prefer period-labeled if available)
    cluster_candidates = [
        f"output/clustering_results_spu_{period_label}.csv",
        CLUSTER_LABELS_FILE,
    ]
    cluster_path = cluster_candidates[0] if os.path.exists(cluster_candidates[0]) else cluster_candidates[1]
    cluster_df = pd.read_csv(cluster_path)
    if 'Cluster' in cluster_df.columns:
        cluster_df = cluster_df.rename(columns={'Cluster': 'cluster_id'})
    # Ensure str_code is string for consistent merging
    if 'str_code' in cluster_df.columns:
        cluster_df['str_code'] = cluster_df['str_code'].astype(str)
    # Ensure cluster_id is int for consistent merging
    if 'cluster_id' in cluster_df.columns:
        cluster_df['cluster_id'] = cluster_df['cluster_id'].astype(int)
    log_progress(f"   ✓ Loaded cluster data: {len(cluster_df):,} stores from {cluster_path}")
    
    # Load product roles (prefer period-labeled if available)
    roles_candidates = [
        f"output/product_role_classifications_{period_label}.csv",
        PRODUCT_ROLES_FILE,
    ]
    roles_path = roles_candidates[0] if os.path.exists(roles_candidates[0]) else roles_candidates[1]
    roles_df = pd.read_csv(roles_path)
    log_progress(f"   ✓ Loaded product roles: {len(roles_df):,} products from {roles_path}")
    
    # Load price bands (prefer period-labeled if available)
    price_candidates = [
        f"output/price_band_analysis_{period_label}.csv",
        PRICE_BANDS_FILE,
    ]
    price_path = price_candidates[0] if os.path.exists(price_candidates[0]) else price_candidates[1]
    price_df = pd.read_csv(price_path)
    log_progress(f"   ✓ Loaded price bands: {len(price_df):,} products from {price_path}")
    
    # Load store attributes (prefer period-labeled if available)
    store_attrs_candidates = [
        f"output/enriched_store_attributes_{period_label}.csv",
        STORE_ATTRIBUTES_FILE,
    ]
    store_attrs_path = store_attrs_candidates[0] if os.path.exists(store_attrs_candidates[0]) else store_attrs_candidates[1]
    store_attrs_df = pd.read_csv(store_attrs_path)
    # Ensure str_code is string for consistent merging
    if 'str_code' in store_attrs_df.columns:
        store_attrs_df['str_code'] = store_attrs_df['str_code'].astype(str)
    log_progress(f"   ✓ Loaded store attributes: {len(store_attrs_df):,} stores from {store_attrs_path}")
    
    return sales_df, cluster_df, roles_df, price_df, store_attrs_df, supply_demand_df

def create_integrated_dataset(sales_df: pd.DataFrame, cluster_df: pd.DataFrame, 
                            roles_df: pd.DataFrame, price_df: pd.DataFrame, 
                            store_attrs_df: pd.DataFrame, supply_demand_df: pd.DataFrame) -> pd.DataFrame:
    """Create integrated dataset with all dimensions"""
    
    # Remove duplicate columns from all input dataframes
    for df_name, df in [("sales_df", sales_df), ("cluster_df", cluster_df), ("roles_df", roles_df), 
                        ("price_df", price_df), ("store_attrs_df", store_attrs_df), ("supply_demand_df", supply_demand_df)]:
        if df.columns.duplicated().any():
            duplicate_cols = df.columns[df.columns.duplicated()].tolist()
            print(f"   ⚠️ Removing duplicate columns from {df_name}: {duplicate_cols}")
            # Keep first occurrence of each column name
            df = df.loc[:, ~df.columns.duplicated(keep='first')]
            
            # Verify critical columns are preserved
            if df_name == "cluster_df" and 'cluster_id' not in df.columns:
                print(f"   ❌ ERROR: cluster_id column was removed from {df_name}!")
                print(f"   📋 Available columns after dedup: {list(df.columns)}")
                raise ValueError(f"cluster_id column missing from {df_name} after duplicate removal")
            
            # Update the reference
            if df_name == "sales_df":
                sales_df = df
            elif df_name == "cluster_df":
                cluster_df = df
            elif df_name == "roles_df":
                roles_df = df
            elif df_name == "price_df":
                price_df = df
            elif df_name == "store_attrs_df":
                store_attrs_df = df
            elif df_name == "supply_demand_df":
                supply_demand_df = df
    log_progress("🔗 Creating integrated dataset for gap analysis...")
    
    # Start with sales data
    integrated_df = sales_df.copy()
    
    # Add cluster information
    # Verify cluster_df has required columns
    if 'cluster_id' not in cluster_df.columns:
        print(f"   ❌ ERROR: cluster_df missing cluster_id column!")
        print(f"   📋 Available cluster_df columns: {list(cluster_df.columns)}")
        raise ValueError("cluster_df missing cluster_id column")
    if 'str_code' not in cluster_df.columns:
        print(f"   ❌ ERROR: cluster_df missing str_code column!")
        print(f"   📋 Available cluster_df columns: {list(cluster_df.columns)}")
        raise ValueError("cluster_df missing str_code column")
        
    integrated_df = integrated_df.merge(cluster_df[['str_code', 'cluster_id']], on='str_code', how='inner')
    
    # Verify cluster_id was added successfully
    if 'cluster_id' not in integrated_df.columns:
        print(f"   ❌ ERROR: cluster_id missing from integrated_df after cluster merge!")
        print(f"   📋 Available integrated_df columns: {list(integrated_df.columns)}")
        raise ValueError("cluster_id missing from integrated_df after cluster merge")
    
    # Add supply-demand gap information if available
    if not supply_demand_df.empty:
        # Merge key supply-demand gap metrics
        # Use available columns from Step 29 output
        available_gap_columns = ['cluster_id']
        gap_columns_to_add = [
            'overall_gap_severity', 'category_diversity', 'subcategory_diversity', 
            'category_concentration', 'dominant_category', 'dominant_category_share',
            'fashion_ratio', 'basic_ratio', 'fashion_gap', 'style_gap_severity',
            'capacity_utilization', 'capacity_pressure'
        ]
        for col in gap_columns_to_add:
            if col in supply_demand_df.columns:
                available_gap_columns.append(col)
        
        if len(available_gap_columns) > 1:  # Only merge if we have additional columns
            gap_metrics = supply_demand_df[available_gap_columns]
            integrated_df = integrated_df.merge(gap_metrics, on='cluster_id', how='left')
    
    # Add product role information
    integrated_df = integrated_df.merge(
        roles_df[['spu_code', 'product_role', 'category', 'subcategory']], 
        on='spu_code', 
        how='left'
    )
    
    # Add price band information
    integrated_df = integrated_df.merge(
        price_df[['spu_code', 'price_band', 'avg_unit_price']], 
        on='spu_code', 
        how='left'
    )
    
    # Add store attributes
    store_capacity = store_attrs_df[['str_code', 'estimated_rack_capacity', 'store_type']].drop_duplicates()
    integrated_df = integrated_df.merge(store_capacity, on='str_code', how='left')
    
    # Calculate derived metrics
    integrated_df['total_sales_amt'] = integrated_df['fashion_sal_amt'] + integrated_df['basic_sal_amt']
    # Use sal_qty as the total quantity column
    if 'sal_qty' in integrated_df.columns:
        integrated_df['total_sales_qty'] = integrated_df['sal_qty']
    else:
        integrated_df['total_sales_qty'] = 0
    integrated_df['fashion_ratio'] = (integrated_df['fashion_sal_amt'] / integrated_df['total_sales_amt']).fillna(0)
    integrated_df['basic_ratio'] = (integrated_df['basic_sal_amt'] / integrated_df['total_sales_amt']).fillna(0)
    
    # Derive seasonal indicator
    integrated_df['seasonal_indicator'] = integrated_df['fashion_ratio'].apply(
        lambda x: 'Seasonal' if x >= 0.7 else 'Year-Round'
    )
    
    log_progress(f"   ✓ Created integrated dataset: {len(integrated_df):,} records")
    
    return integrated_df

# ===== 6-DIMENSIONAL COVERAGE ANALYSIS =====

def analyze_cluster_coverage(integrated_df: pd.DataFrame, cluster_id: int) -> Dict[str, Any]:
    """Analyze coverage across all 6 dimensions for a cluster"""
    
    cluster_data = integrated_df[integrated_df['cluster_id'] == cluster_id]
    stores_in_cluster = cluster_data['str_code'].unique()
    
    coverage_analysis = {
        'cluster_id': cluster_id,
        'store_count': len(stores_in_cluster),
        'product_count': len(cluster_data),
        'total_sales_amt': cluster_data['total_sales_amt'].sum()
    }
    
    # 1. Category Coverage
    category_count = cluster_data['category'].nunique() if 'category' in cluster_data.columns else 0
    subcategory_count = cluster_data['subcategory'].nunique() if 'subcategory' in cluster_data.columns else 0
    category_coverage_score = min(category_count / 10, 1.0)  # Normalized to 0-1
    
    coverage_analysis['category_coverage'] = {
        'category_count': category_count,
        'subcategory_count': subcategory_count,
        'coverage_score': category_coverage_score,
        'status': 'optimal' if 7 <= category_count <= 12 else 'adequate' if category_count >= 5 else 'insufficient'
    }
    
    # 2. Price Band Coverage
    price_bands = cluster_data['price_band'].value_counts()
    price_band_count = len(price_bands)
    price_band_balance = 1 - ((price_bands / len(cluster_data))**2).sum()  # Herfindahl index
    
    coverage_analysis['price_band_coverage'] = {
        'price_band_count': price_band_count,
        'price_band_balance': price_band_balance,
        'price_distribution': price_bands.to_dict(),
        'status': 'optimal' if 4 <= price_band_count <= 5 else 'adequate' if price_band_count >= 3 else 'insufficient'
    }
    
    # 3. Style Orientation
    avg_fashion_ratio = cluster_data['fashion_ratio'].mean()
    style_balance_score = 1 - abs(avg_fashion_ratio - 0.5) * 2  # Closer to 0.5 = better balance
    
    coverage_analysis['style_orientation'] = {
        'fashion_ratio': avg_fashion_ratio,
        'basic_ratio': 1 - avg_fashion_ratio,
        'balance_score': style_balance_score,
        'status': 'optimal' if 0.4 <= avg_fashion_ratio <= 0.6 else 'adequate' if 0.3 <= avg_fashion_ratio <= 0.7 else 'imbalanced'
    }
    
    # 4. Product Role Balance
    role_distribution = cluster_data['product_role'].value_counts(normalize=True)
    role_count = len(role_distribution)
    role_diversity_score = 1 - (role_distribution**2).sum()  # Diversity index
    
    coverage_analysis['product_role_balance'] = {
        'role_count': role_count,
        'role_distribution': role_distribution.to_dict(),
        'diversity_score': role_diversity_score,
        'status': 'optimal' if role_count == 4 else 'adequate' if role_count >= 3 else 'insufficient'
    }
    
    # 5. Seasonal Responsiveness
    seasonal_distribution = cluster_data['seasonal_indicator'].value_counts(normalize=True)
    seasonal_ratio = seasonal_distribution.get('Seasonal', 0)
    
    coverage_analysis['seasonal_responsiveness'] = {
        'seasonal_ratio': seasonal_ratio,
        'year_round_ratio': 1 - seasonal_ratio,
        'seasonal_distribution': seasonal_distribution.to_dict(),
        'status': 'optimal' if 0.25 <= seasonal_ratio <= 0.4 else 'adequate' if seasonal_ratio >= 0.2 else 'insufficient'
    }
    
    # 6. Capacity Utilization
    avg_capacity = cluster_data['estimated_rack_capacity'].mean()
    products_per_store = len(cluster_data) / len(stores_in_cluster)
    capacity_utilization = products_per_store / avg_capacity if avg_capacity > 0 else 0
    
    coverage_analysis['capacity_utilization'] = {
        'avg_capacity': avg_capacity,
        'products_per_store': products_per_store,
        'utilization_rate': capacity_utilization,
        'status': 'optimal' if 0.6 <= capacity_utilization <= 0.8 else 'adequate' if capacity_utilization >= 0.4 else 'underutilized' if capacity_utilization < 0.4 else 'overutilized'
    }
    
    # Overall coverage score
    dimension_scores = [
        coverage_analysis['category_coverage']['coverage_score'],
        coverage_analysis['price_band_coverage']['price_band_balance'],
        coverage_analysis['style_orientation']['balance_score'],
        coverage_analysis['product_role_balance']['diversity_score'],
        coverage_analysis['seasonal_responsiveness']['seasonal_ratio'] * 2.5,  # Scale to 0-1
        min(coverage_analysis['capacity_utilization']['utilization_rate'] / 0.7, 1.0)  # Scale to 0-1
    ]
    
    overall_coverage_score = np.mean(dimension_scores)
    coverage_analysis['overall_coverage_score'] = overall_coverage_score
    coverage_analysis['overall_status'] = 'excellent' if overall_coverage_score >= 0.8 else 'good' if overall_coverage_score >= 0.6 else 'needs_improvement'
    
    return coverage_analysis

def create_coverage_matrix(integrated_df: pd.DataFrame) -> pd.DataFrame:
    """Create comprehensive coverage matrix across all clusters and dimensions"""
    log_progress("📋 Creating 6-dimensional coverage matrix...")
    
    clusters = sorted(integrated_df['cluster_id'].unique())
    coverage_data = []
    
    for cluster_id in clusters:
        log_progress(f"   🔍 Analyzing Cluster {cluster_id} coverage...")
        coverage_analysis = analyze_cluster_coverage(integrated_df, cluster_id)
        
        # Flatten data for matrix
        row_data = {
            'cluster_id': cluster_id,
            'store_count': coverage_analysis['store_count'],
            'product_count': coverage_analysis['product_count'],
            'total_sales_amt': coverage_analysis['total_sales_amt'],
            'overall_coverage_score': coverage_analysis['overall_coverage_score'],
            'overall_status': coverage_analysis['overall_status'],
            
            # Dimension 1: Category Coverage
            'category_count': coverage_analysis['category_coverage']['category_count'],
            'subcategory_count': coverage_analysis['category_coverage']['subcategory_count'],
            'category_coverage_score': coverage_analysis['category_coverage']['coverage_score'],
            'category_status': coverage_analysis['category_coverage']['status'],
            
            # Dimension 2: Price Band Coverage
            'price_band_count': coverage_analysis['price_band_coverage']['price_band_count'],
            'price_band_balance': coverage_analysis['price_band_coverage']['price_band_balance'],
            'price_band_status': coverage_analysis['price_band_coverage']['status'],
            
            # Dimension 3: Style Orientation
            'fashion_ratio': coverage_analysis['style_orientation']['fashion_ratio'],
            'style_balance_score': coverage_analysis['style_orientation']['balance_score'],
            'style_status': coverage_analysis['style_orientation']['status'],
            
            # Dimension 4: Product Role Balance
            'role_count': coverage_analysis['product_role_balance']['role_count'],
            'role_diversity_score': coverage_analysis['product_role_balance']['diversity_score'],
            'role_status': coverage_analysis['product_role_balance']['status'],
            
            # Dimension 5: Seasonal Responsiveness
            'seasonal_ratio': coverage_analysis['seasonal_responsiveness']['seasonal_ratio'],
            'seasonal_status': coverage_analysis['seasonal_responsiveness']['status'],
            
            # Dimension 6: Capacity Utilization
            'capacity_utilization_rate': coverage_analysis['capacity_utilization']['utilization_rate'],
            'capacity_status': coverage_analysis['capacity_utilization']['status']
        }
        
        coverage_data.append(row_data)
    
    coverage_matrix_df = pd.DataFrame(coverage_data)
    
    log_progress(f"   ✓ Coverage matrix created: {len(coverage_matrix_df)} clusters analyzed")
    
    return coverage_matrix_df

# ===== EXECUTIVE SUMMARY GENERATION =====

def create_executive_summary(coverage_matrix_df: pd.DataFrame, integrated_df: pd.DataFrame) -> Dict[str, Any]:
    """Create executive summary of gap analysis findings"""
    log_progress("📊 Creating executive summary...")
    
    total_clusters = len(coverage_matrix_df)
    total_stores = coverage_matrix_df['store_count'].sum()
    total_products = coverage_matrix_df['product_count'].sum()
    total_sales = coverage_matrix_df['total_sales_amt'].sum()
    
    # Overall performance assessment
    excellent_clusters = len(coverage_matrix_df[coverage_matrix_df['overall_status'] == 'excellent'])
    good_clusters = len(coverage_matrix_df[coverage_matrix_df['overall_status'] == 'good'])
    needs_improvement_clusters = len(coverage_matrix_df[coverage_matrix_df['overall_status'] == 'needs_improvement'])
    
    # Dimension-specific analysis
    dimension_summary = {}
    
    # Category coverage summary
    avg_category_count = coverage_matrix_df['category_count'].mean()
    optimal_category_clusters = len(coverage_matrix_df[coverage_matrix_df['category_status'] == 'optimal'])
    
    dimension_summary['category_coverage'] = {
        'avg_categories_per_cluster': avg_category_count,
        'optimal_clusters': optimal_category_clusters,
        'clusters_needing_improvement': total_clusters - optimal_category_clusters,
        'dimension_health': 'good' if optimal_category_clusters / total_clusters >= 0.6 else 'needs_attention'
    }
    
    # Price band coverage summary
    avg_price_band_balance = coverage_matrix_df['price_band_balance'].mean()
    optimal_price_clusters = len(coverage_matrix_df[coverage_matrix_df['price_band_status'] == 'optimal'])
    
    dimension_summary['price_band_coverage'] = {
        'avg_price_balance_score': avg_price_band_balance,
        'optimal_clusters': optimal_price_clusters,
        'clusters_needing_improvement': total_clusters - optimal_price_clusters,
        'dimension_health': 'good' if optimal_price_clusters / total_clusters >= 0.6 else 'needs_attention'
    }
    
    # Style orientation summary
    avg_fashion_ratio = coverage_matrix_df['fashion_ratio'].mean()
    balanced_style_clusters = len(coverage_matrix_df[coverage_matrix_df['style_status'] == 'optimal'])
    
    dimension_summary['style_orientation'] = {
        'avg_fashion_ratio': avg_fashion_ratio,
        'balanced_clusters': balanced_style_clusters,
        'clusters_needing_rebalancing': total_clusters - balanced_style_clusters,
        'dimension_health': 'good' if balanced_style_clusters / total_clusters >= 0.6 else 'needs_attention'
    }
    
    # Product role balance summary
    avg_role_diversity = coverage_matrix_df['role_diversity_score'].mean()
    optimal_role_clusters = len(coverage_matrix_df[coverage_matrix_df['role_status'] == 'optimal'])
    
    dimension_summary['product_role_balance'] = {
        'avg_role_diversity_score': avg_role_diversity,
        'optimal_clusters': optimal_role_clusters,
        'clusters_needing_improvement': total_clusters - optimal_role_clusters,
        'dimension_health': 'good' if optimal_role_clusters / total_clusters >= 0.6 else 'needs_attention'
    }
    
    # Seasonal responsiveness summary
    avg_seasonal_ratio = coverage_matrix_df['seasonal_ratio'].mean()
    optimal_seasonal_clusters = len(coverage_matrix_df[coverage_matrix_df['seasonal_status'] == 'optimal'])
    
    dimension_summary['seasonal_responsiveness'] = {
        'avg_seasonal_ratio': avg_seasonal_ratio,
        'optimal_clusters': optimal_seasonal_clusters,
        'clusters_needing_improvement': total_clusters - optimal_seasonal_clusters,
        'dimension_health': 'good' if optimal_seasonal_clusters / total_clusters >= 0.6 else 'needs_attention'
    }
    
    # Capacity utilization summary
    avg_capacity_utilization = coverage_matrix_df['capacity_utilization_rate'].mean()
    optimal_capacity_clusters = len(coverage_matrix_df[coverage_matrix_df['capacity_status'] == 'optimal'])
    
    dimension_summary['capacity_utilization'] = {
        'avg_capacity_utilization': avg_capacity_utilization,
        'optimal_clusters': optimal_capacity_clusters,
        'clusters_needing_optimization': total_clusters - optimal_capacity_clusters,
        'dimension_health': 'good' if optimal_capacity_clusters / total_clusters >= 0.6 else 'needs_attention'
    }
    
    # Key insights and recommendations
    key_insights = []
    recommendations = []
    
    # Identify top issues
    dimension_health_scores = {dim: data['dimension_health'] for dim, data in dimension_summary.items()}
    needs_attention_dimensions = [dim for dim, health in dimension_health_scores.items() if health == 'needs_attention']
    
    if needs_attention_dimensions:
        key_insights.append(f"Critical gaps identified in {len(needs_attention_dimensions)} dimensions: {', '.join(needs_attention_dimensions)}")
        recommendations.append(f"Priority focus areas: {', '.join(needs_attention_dimensions)}")
    
    # Cluster performance insights
    if excellent_clusters == 0:
        key_insights.append("No clusters achieving excellent coverage across all dimensions")
        recommendations.append("Implement comprehensive optimization across all clusters")
    elif excellent_clusters < total_clusters / 2:
        key_insights.append(f"Only {excellent_clusters}/{total_clusters} clusters achieving excellent performance")
        recommendations.append("Focus on elevating good-performing clusters to excellent status")
    
    # Capacity insights
    if avg_capacity_utilization < 0.5:
        key_insights.append("Low average capacity utilization indicates potential for expansion")
        recommendations.append("Consider increasing product allocation where capacity allows")
    elif avg_capacity_utilization > 0.8:
        key_insights.append("High capacity utilization may limit growth opportunities")
        recommendations.append("Optimize space efficiency and consider capacity expansion")
    
    executive_summary = {
        'analysis_metadata': {
            'total_clusters_analyzed': total_clusters,
            'total_stores_covered': total_stores,
            'total_products_analyzed': total_products,
            'total_sales_amt': total_sales,
            'analysis_timestamp': datetime.now().isoformat()
        },
        'overall_performance': {
            'excellent_clusters': excellent_clusters,
            'good_clusters': good_clusters,
            'needs_improvement_clusters': needs_improvement_clusters,
            'overall_health_score': (excellent_clusters * 3 + good_clusters * 2 + needs_improvement_clusters * 1) / (total_clusters * 3)
        },
        'dimension_summary': dimension_summary,
        'key_insights': key_insights,
        'recommendations': recommendations,
        'priority_actions': [
            "Address critical gaps in underperforming dimensions",
            "Optimize capacity utilization across all clusters", 
            "Enhance product role balance and seasonal responsiveness",
            "Implement cluster-specific improvement strategies"
        ]
    }
    
    return executive_summary

# ===== EXCEL WORKBOOK CREATION =====

def create_gap_analysis_workbook(coverage_matrix_df: pd.DataFrame, executive_summary: Dict[str, Any], 
                               integrated_df: pd.DataFrame, output_filename: str) -> None:
    """Create comprehensive Excel workbook with multiple sheets"""
    log_progress("📊 Creating comprehensive gap analysis workbook...")
    
    if not EXCEL_AVAILABLE:
        log_progress("⚠️ Excel formatting not available - creating CSV outputs instead")
        create_csv_outputs(coverage_matrix_df, executive_summary, integrated_df, 
                          CLUSTER_COVERAGE_MATRIX, GAP_WORKBOOK_CSV, EXECUTIVE_SUMMARY_JSON, 
                          "output/cluster_coverage_matrix.csv", "output/gap_analysis_workbook_data.csv", "output/gap_workbook_executive_summary.json")
        return
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Sheet 1: Executive Summary
    create_executive_summary_sheet(wb, executive_summary)
    
    # Sheet 2: Coverage Matrix
    create_coverage_matrix_sheet(wb, coverage_matrix_df)
    
    # Sheet 3: Cluster Details
    create_cluster_details_sheet(wb, coverage_matrix_df, integrated_df)
    
    # Sheet 4: Store-Level Data
    create_store_level_sheet(wb, integrated_df)
    
    # Sheet 5: Action Plan
    create_action_plan_sheet(wb, coverage_matrix_df, executive_summary)
    
    # Save workbook
    wb.save(output_filename)
    log_progress(f"✅ Created gap analysis workbook: {output_filename}")

def create_executive_summary_sheet(wb: openpyxl.Workbook, executive_summary: Dict[str, Any]) -> None:
    """Create executive summary sheet"""
    ws = wb.create_sheet("Executive Summary", 0)
    
    # Title
    ws['A1'] = "Gap Analysis Executive Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # Metadata
    metadata = executive_summary['analysis_metadata']
    ws['A3'] = "Analysis Overview"
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = f"Clusters Analyzed: {metadata['total_clusters_analyzed']}"
    ws['A5'] = f"Stores Covered: {metadata['total_stores_covered']}"
    ws['A6'] = f"Products Analyzed: {metadata['total_products_analyzed']:,}"
    ws['A7'] = f"Total Sales: ¥{metadata['total_sales_amt']:,.0f}"
    ws['A8'] = f"Analysis Date: {metadata['analysis_timestamp'][:10]}"
    
    # Overall Performance
    performance = executive_summary['overall_performance']
    ws['A10'] = "Overall Performance"
    ws['A10'].font = Font(bold=True)
    
    ws['A11'] = f"Excellent Clusters: {performance['excellent_clusters']}"
    ws['A12'] = f"Good Clusters: {performance['good_clusters']}"
    ws['A13'] = f"Needs Improvement: {performance['needs_improvement_clusters']}"
    ws['A14'] = f"Health Score: {performance['overall_health_score']:.1%}"
    
    # Key Insights
    ws['A16'] = "Key Insights"
    ws['A16'].font = Font(bold=True)
    
    for i, insight in enumerate(executive_summary['key_insights']):
        ws[f'A{17+i}'] = f"• {insight}"
    
    # Recommendations
    row_start = 17 + len(executive_summary['key_insights']) + 2
    ws[f'A{row_start}'] = "Recommendations"
    ws[f'A{row_start}'].font = Font(bold=True)
    
    for i, recommendation in enumerate(executive_summary['recommendations']):
        ws[f'A{row_start+1+i}'] = f"• {recommendation}"

def create_coverage_matrix_sheet(wb: openpyxl.Workbook, coverage_matrix_df: pd.DataFrame) -> None:
    """Create coverage matrix sheet with conditional formatting"""
    ws = wb.create_sheet("Coverage Matrix")
    
    # Add title
    ws['A1'] = "6-Dimensional Coverage Matrix"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Add headers
    headers = list(coverage_matrix_df.columns)
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Add data
    for row, data_row in enumerate(coverage_matrix_df.itertuples(index=False), 4):
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Apply conditional formatting for status columns
    status_columns = ['overall_status', 'category_status', 'price_band_status', 'style_status', 'role_status', 'seasonal_status', 'capacity_status']
    
    for col_name in status_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # Green for optimal/excellent
            ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{len(coverage_matrix_df)+3}',
                CellIsRule(operator='containsText', formula=['"optimal"'], 
                          fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
            ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{len(coverage_matrix_df)+3}',
                CellIsRule(operator='containsText', formula=['"excellent"'], 
                          fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")))
            
            # Yellow for adequate/good
            ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{len(coverage_matrix_df)+3}',
                CellIsRule(operator='containsText', formula=['"adequate"'], 
                          fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))
            ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{len(coverage_matrix_df)+3}',
                CellIsRule(operator='containsText', formula=['"good"'], 
                          fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")))
            
            # Red for insufficient/needs_improvement
            ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{len(coverage_matrix_df)+3}',
                CellIsRule(operator='containsText', formula=['"insufficient"'], 
                          fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))
            ws.conditional_formatting.add(f'{col_letter}4:{col_letter}{len(coverage_matrix_df)+3}',
                CellIsRule(operator='containsText', formula=['"needs_improvement"'], 
                          fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")))

def create_cluster_details_sheet(wb: openpyxl.Workbook, coverage_matrix_df: pd.DataFrame, integrated_df: pd.DataFrame) -> None:
    """Create detailed cluster analysis sheet"""
    ws = wb.create_sheet("Cluster Details")
    
    ws['A1'] = "Detailed Cluster Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    row = 3
    for _, cluster_data in coverage_matrix_df.iterrows():
        cluster_id = cluster_data['cluster_id']
        
        # Cluster header
        ws[f'A{row}'] = f"Cluster {cluster_id}"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        row += 1
        
        # Cluster metrics
        ws[f'A{row}'] = f"Stores: {cluster_data['store_count']}"
        ws[f'B{row}'] = f"Products: {cluster_data['product_count']}"
        ws[f'C{row}'] = f"Sales: ¥{cluster_data['total_sales_amt']:,.0f}"
        ws[f'D{row}'] = f"Overall Score: {cluster_data['overall_coverage_score']:.1%}"
        
        row += 2

def create_store_level_sheet(wb: openpyxl.Workbook, integrated_df: pd.DataFrame) -> None:
    """Create store-level disaggregated data sheet"""
    ws = wb.create_sheet("Store Level Data")
    
    # Aggregate to store level
    store_level_data = integrated_df.groupby(['str_code', 'cluster_id']).agg({
        'total_sales_amt': 'sum',
        'total_sales_qty': 'sum',
        'category': 'nunique',
        'subcategory': 'nunique',
        'product_role': 'nunique',
        'price_band': 'nunique',
        'fashion_ratio': 'mean',
        'estimated_rack_capacity': 'first',
        'store_type': 'first'
    }).reset_index()
    
    store_level_data.columns = ['Store_Code', 'Cluster_ID', 'Total_Sales_Amt', 'Total_Sales_Qty', 
                               'Category_Count', 'Subcategory_Count', 'Role_Count', 'Price_Band_Count',
                               'Fashion_Ratio', 'Estimated_Capacity', 'Store_Type']
    
    # Add calculated metrics
    store_level_data['Products_Count'] = integrated_df.groupby('str_code').size().values
    store_level_data['Capacity_Utilization'] = store_level_data['Products_Count'] / store_level_data['Estimated_Capacity']
    store_level_data['Sales_Per_Product'] = store_level_data['Total_Sales_Amt'] / store_level_data['Products_Count']
    
    # Add constraint status
    def get_constraint_status(row):
        if row['Capacity_Utilization'] > 0.9:
            return 'Over-Capacity'
        elif row['Capacity_Utilization'] < 0.3:
            return 'Under-Utilized'
        elif row['Category_Count'] < 3:
            return 'Low-Diversity'
        else:
            return 'Normal'
    
    store_level_data['Constraint_Status'] = store_level_data.apply(get_constraint_status, axis=1)
    
    # Add to worksheet
    ws['A1'] = "Store-Level Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Headers
    for col, header in enumerate(store_level_data.columns, 1):
        ws.cell(row=3, column=col, value=header)
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Data
    for row, data_row in enumerate(store_level_data.itertuples(index=False), 4):
        for col, value in enumerate(data_row, 1):
            ws.cell(row=row, column=col, value=value)

def create_action_plan_sheet(wb: openpyxl.Workbook, coverage_matrix_df: pd.DataFrame, executive_summary: Dict[str, Any]) -> None:
    """Create action plan sheet"""
    ws = wb.create_sheet("Action Plan")
    
    ws['A1'] = "Gap Analysis Action Plan"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Priority actions from executive summary
    ws['A3'] = "Priority Actions"
    ws['A3'].font = Font(bold=True)
    
    for i, action in enumerate(executive_summary['priority_actions']):
        ws[f'A{4+i}'] = f"{i+1}. {action}"
    
    # Cluster-specific recommendations
    row = 4 + len(executive_summary['priority_actions']) + 2
    ws[f'A{row}'] = "Cluster-Specific Recommendations"
    ws[f'A{row}'].font = Font(bold=True)
    
    row += 1
    for _, cluster_data in coverage_matrix_df.iterrows():
        cluster_id = cluster_data['cluster_id']
        ws[f'A{row}'] = f"Cluster {cluster_id}:"
        
        recommendations = []
        if cluster_data['category_status'] != 'optimal':
            recommendations.append("Expand category coverage")
        if cluster_data['price_band_status'] != 'optimal':
            recommendations.append("Rebalance price bands")
        if cluster_data['style_status'] != 'optimal':
            recommendations.append("Adjust fashion/basic mix")
        if cluster_data['role_status'] != 'optimal':
            recommendations.append("Optimize product roles")
        if cluster_data['seasonal_status'] != 'optimal':
            recommendations.append("Enhance seasonal responsiveness")
        if cluster_data['capacity_status'] not in ['optimal', 'adequate']:
            recommendations.append("Optimize capacity utilization")
        
        if recommendations:
            ws[f'B{row}'] = "; ".join(recommendations)
        else:
            ws[f'B{row}'] = "Maintain current performance"
        
        row += 1

def create_csv_outputs(coverage_matrix_df: pd.DataFrame, executive_summary: Dict[str, Any], integrated_df: pd.DataFrame,
                      coverage_matrix_filename: str, workbook_data_filename: str, summary_filename: str, 
                      generic_coverage_matrix: str, generic_workbook_csv: str, generic_summary_json: str) -> None:
    """Create CSV outputs when Excel is not available"""
    log_progress("📄 Creating CSV outputs...")
    
    # Save coverage matrix (DUAL OUTPUT PATTERN)
    # Save timestamped version (for backup/inspection)
    coverage_matrix_df.to_csv(coverage_matrix_filename, index=False)
    log_progress(f"   ✓ Saved timestamped coverage matrix: {coverage_matrix_filename}")
    
    # Create symlink for generic version (for pipeline flow)
    if os.path.exists(generic_coverage_matrix) or os.path.islink(generic_coverage_matrix):
        os.remove(generic_coverage_matrix)
    os.symlink(os.path.basename(coverage_matrix_filename), generic_coverage_matrix)
    log_progress(f"   ✓ Created symlink: {generic_coverage_matrix} -> {coverage_matrix_filename}")
    
    # Save main workbook data (DUAL OUTPUT PATTERN)
    # Save timestamped version (for backup/inspection)
    coverage_matrix_df.to_csv(workbook_data_filename, index=False)
    log_progress(f"   ✓ Saved timestamped workbook data: {workbook_data_filename}")
    
    # Create symlink for generic version (for pipeline flow)
    if os.path.exists(generic_workbook_csv) or os.path.islink(generic_workbook_csv):
        os.remove(generic_workbook_csv)
    os.symlink(os.path.basename(workbook_data_filename), generic_workbook_csv)
    log_progress(f"   ✓ Created symlink: {generic_workbook_csv} -> {workbook_data_filename}")
    
    # Save executive summary (DUAL OUTPUT PATTERN)
    # Save timestamped version (for backup/inspection)
    with open(summary_filename, 'w') as f:
        json.dump(executive_summary, f, indent=2, default=str)
    log_progress(f"   ✓ Saved timestamped executive summary: {summary_filename}")
    
    # Save generic version (for pipeline flow)
    with open(generic_summary_json, 'w') as f:
        json.dump(executive_summary, f, indent=2, default=str)
    log_progress(f"   ✓ Saved generic executive summary: {generic_summary_json}")

# ===== MAIN EXECUTION =====

def main() -> None:
    """Main function for gap analysis workbook generation (period-aware)"""
    start_time = datetime.now()
    log_progress("🚀 Starting TOP PRIORITY: Gap-Analysis Workbook Generation...")
    
    # Parse CLI and derive period label
    args = _parse_args()
    target_yyyymm = args.target_yyyymm
    target_period = args.target_period
    period_label = get_period_label(target_yyyymm, target_period)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # DUAL OUTPUT PATTERN - Define both timestamped and generic versions
    # Timestamped versions (for backup/inspection)
    timestamped_workbook_excel = f"output/gap_analysis_workbook_{period_label}_{timestamp}.xlsx"
    timestamped_workbook_csv = f"output/gap_analysis_workbook_data_{period_label}_{timestamp}.csv"
    timestamped_summary_json = f"output/gap_workbook_executive_summary_{period_label}_{timestamp}.json"
    timestamped_coverage_matrix = f"output/cluster_coverage_matrix_{period_label}_{timestamp}.csv"
    
    # Generic versions (for pipeline flow)
    generic_workbook_excel = "output/gap_analysis_workbook.xlsx"
    generic_workbook_csv = "output/gap_analysis_workbook_data.csv"
    generic_summary_json = "output/gap_workbook_executive_summary.json"
    generic_coverage_matrix = "output/cluster_coverage_matrix.csv"
    
    # Use timestamped versions for manifest registration
    GAP_WORKBOOK_EXCEL = timestamped_workbook_excel
    GAP_WORKBOOK_CSV = timestamped_workbook_csv
    EXECUTIVE_SUMMARY_JSON = timestamped_summary_json
    CLUSTER_COVERAGE_MATRIX = timestamped_coverage_matrix
    
    log_progress(f"   Period: {period_label}")
    log_progress(f"   Timestamp: {timestamp}")
    
    try:
        # Load data
        sales_df, cluster_df, roles_df, price_df, store_attrs_df, supply_demand_df = load_workbook_data(target_yyyymm, target_period, period_label)
        
        # Create integrated dataset
        integrated_df = create_integrated_dataset(sales_df, cluster_df, roles_df, price_df, store_attrs_df, supply_demand_df)
        
        # Create coverage matrix
        coverage_matrix_df = create_coverage_matrix(integrated_df)
        
        # Create executive summary
        executive_summary = create_executive_summary(coverage_matrix_df, integrated_df)
        
        # Create comprehensive workbook
        create_gap_analysis_workbook(coverage_matrix_df, executive_summary, integrated_df, GAP_WORKBOOK_EXCEL)
        
        # Create CSV outputs as well
        create_csv_outputs(coverage_matrix_df, executive_summary, integrated_df, 
                          CLUSTER_COVERAGE_MATRIX, GAP_WORKBOOK_CSV, EXECUTIVE_SUMMARY_JSON, 
                          generic_coverage_matrix, generic_workbook_csv, generic_summary_json)
        
        # Summary results
        log_progress("\n🎯 GAP ANALYSIS WORKBOOK RESULTS:")
        log_progress(f"   📊 Clusters Analyzed: {len(coverage_matrix_df)}")
        log_progress(f"   🏪 Stores Covered: {coverage_matrix_df['store_count'].sum()}")
        log_progress(f"   📦 Products Analyzed: {coverage_matrix_df['product_count'].sum():,}")
        log_progress(f"   💰 Total Sales: ¥{coverage_matrix_df['total_sales_amt'].sum():,.0f}")
        
        excellent_clusters = len(coverage_matrix_df[coverage_matrix_df['overall_status'] == 'excellent'])
        good_clusters = len(coverage_matrix_df[coverage_matrix_df['overall_status'] == 'good'])
        log_progress(f"   ⭐ Excellent Clusters: {excellent_clusters}")
        log_progress(f"   ✅ Good Clusters: {good_clusters}")
        log_progress(f"   📈 Overall Health Score: {executive_summary['overall_performance']['overall_health_score']:.1%}")
        
        log_progress(f"\n✅ Gap Analysis Workbook completed in {(datetime.now() - start_time).total_seconds():.1f} seconds")
        
        log_progress(f"\n📁 Generated Files:")
        log_progress(f"   • {GAP_WORKBOOK_EXCEL if EXCEL_AVAILABLE else GAP_WORKBOOK_CSV}")
        log_progress(f"   • {CLUSTER_COVERAGE_MATRIX}")
        log_progress(f"   • {EXECUTIVE_SUMMARY_JSON}")
        
        # Register outputs in pipeline manifest (generic + period-specific)
        try:
            records = len(coverage_matrix_df)
            columns = len(coverage_matrix_df.columns) if not coverage_matrix_df.empty else 0
        except Exception:
            records, columns = 0, 0
        
        metadata = {
            "target_year": int(target_yyyymm[:4]),
            "target_month": int(target_yyyymm[4:]),
            "target_period": target_period,
            "records": records,
            "columns": columns
        }
        
        register_step_output("step31", "gap_analysis_workbook_excel", GAP_WORKBOOK_EXCEL, metadata)
        register_step_output("step31", f"gap_analysis_workbook_excel_{period_label}", GAP_WORKBOOK_EXCEL, metadata)
        register_step_output("step31", "gap_analysis_workbook_csv", GAP_WORKBOOK_CSV, metadata)
        register_step_output("step31", f"gap_analysis_workbook_csv_{period_label}", GAP_WORKBOOK_CSV, metadata)
        register_step_output("step31", "gap_workbook_executive_summary", EXECUTIVE_SUMMARY_JSON, metadata)
        register_step_output("step31", f"gap_workbook_executive_summary_{period_label}", EXECUTIVE_SUMMARY_JSON, metadata)
        register_step_output("step31", "cluster_coverage_matrix", CLUSTER_COVERAGE_MATRIX, metadata)
        register_step_output("step31", f"cluster_coverage_matrix_{period_label}", CLUSTER_COVERAGE_MATRIX, metadata)
        log_progress("📋 Manifest updated with Step 31 outputs (generic + period-specific)")
        
    except Exception as e:
        log_progress(f"❌ Error in gap analysis workbook generation: {e}")
        raise

if __name__ == "__main__":
    main() 