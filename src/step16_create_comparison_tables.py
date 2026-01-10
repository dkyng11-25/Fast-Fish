#!/usr/bin/env python3
"""
Step 16: Create Comparison Tables
================================

Generate Excel-compatible comparison tables for:
- Period-aware Year-over-Year (YOY) baseline vs current period
- Category and Store Group comparisons using Step 15 outputs
- Performance benchmarks for Fast Fish analysis

Pipeline Flow:
Step 15 → Step 16 → Step 17 → Step 18

 HOW TO RUN (CLI + ENV) — Read this before executing
 ----------------------------------------------------
 Overview
 - Period-aware: Step 16 builds Excel-friendly comparison tables using Step 15 outputs for a given target period and its YoY baseline.
 - Why Step 15 must match: the YOY and historical files are tagged by the baseline label (e.g., 202410A for target 202510A). Using a mismatched baseline causes wrong-period comparisons.

 Quick Start (target 202510A → baseline 202410A by default)
   Command:
     PYTHONPATH=. python3 src/step16_create_comparison_tables.py \
       --target-yyyymm 202510 \
       --target-period A

 Inputs and Overrides
 - Default resolution: this step resolves Step 15 outputs via the manifest for the exact baseline label. If not found, it falls back to timestamped files matching `historical_reference_{baseline}_*.csv` and `year_over_year_comparison_{baseline}_*.csv`.
 - --yoy-file / --historical-file: use these to provide explicit Step 15 outputs when the manifest isn’t available or when running ad hoc.
 - --baseline-yyyymm / --baseline-period: override default YoY baseline if business requires it (document rationale to avoid seasonal skew).

 Why this configuration leads to stable outcomes
 - Using the exact Step 15 baseline ensures table metrics compare like-for-like seasonal windows.
 - Manifest-driven resolution avoids accidentally mixing old or wrong-period files and keeps downstream consumption deterministic.

 Common failure modes (and what to do)
 - "Step 15 YOY comparison not found for {baseline_label}"
   • Cause: Step 15 didn’t register outputs for that baseline, or files aren’t present on disk.
   • Fix: rerun Step 15 or pass --yoy-file/--historical-file explicitly. Verify the baseline label (e.g., 202410A).
 - Empty/low-coverage tables
   • Cause: category/subcategory naming differences between historical vs. current, or missing columns in source files.
   • Fix: ensure Step 15 created expected columns and consistent taxonomy; confirm Step 14/15 pipeline used the same taxonomy normalization.
 - Excel writer not available
   • Cause: missing `openpyxl` when trying to save .xlsx outputs in your own extensions.
   • Fix: `pip install openpyxl` or write CSV outputs; this script already focuses on table creation, you can export downstream.

 Manifest notes
 - This step reads Step 15 outputs and registers its own comparison tables. Downstream steps (17–18) should use manifest lookups to avoid hardcoding paths.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import logging
import argparse
from typing import Dict, List, Tuple, Optional
from pipeline_manifest import get_manifest, register_step_output
from config import get_api_data_files
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def _parse_args():
    """Parse CLI arguments for period-aware Step 16."""
    parser = argparse.ArgumentParser(description="Step 16: Create comparison tables (period-aware)")
    parser.add_argument("--target-yyyymm", required=True, help="Target year-month for current run, e.g. 202509")
    parser.add_argument("--target-period", choices=["A", "B"], required=True, help="Target period (A or B)")
    parser.add_argument("--baseline-yyyymm", help="Override baseline year-month (defaults to last-year same month)")
    parser.add_argument("--baseline-period", choices=["A", "B"], help="Override baseline period (defaults to target-period)")
    parser.add_argument("--yoy-file", help="Path to Step 15 YOY comparison CSV (optional, overrides manifest)")
    parser.add_argument("--historical-file", help="Path to Step 15 historical reference CSV (optional, overrides manifest)")
    parser.add_argument("--historical-raw-file", help="Path to baseline raw SPU sales CSV (optional)")
    parser.add_argument("--current-raw-file", help="Path to current raw SPU sales CSV (optional)")
    parser.add_argument(
        "--focus-keywords",
        help="DEPRECATED (ignored): Use --fashion-keywords and --basic-keywords to control Fashion vs Basic segmentation.",
    )
    parser.add_argument(
        "--fashion-keywords",
        help="Comma-separated keywords defining the 'Fashion' segment (zh/en). Overrides/extends defaults.",
    )
    parser.add_argument(
        "--basic-keywords",
        help="Comma-separated keywords defining the 'Basic' segment (zh/en). Overrides/extends defaults.",
    )
    parser.add_argument(
        "--hybrid-threshold",
        type=float,
        default=0.15,
        help="Threshold around 50%% fashion share to label as Hybrid (e.g., 0.15 => 35%%–65%% is Hybrid)",
    )
    return parser.parse_args()

def _compute_baseline_yyyymm(target_yyyymm: str) -> str:
    if not (isinstance(target_yyyymm, str) and len(target_yyyymm) == 6 and target_yyyymm.isdigit()):
        raise ValueError(f"Invalid target_yyyymm: {target_yyyymm}")
    year = int(target_yyyymm[:4])
    month = int(target_yyyymm[4:6])
    return f"{year - 1}{month:02d}"

def load_data(
    target_yyyymm: str,
    target_period: str,
    baseline_yyyymm: Optional[str] = None,
    baseline_period: Optional[str] = None,
    yoy_file_override: Optional[str] = None,
    historical_file_override: Optional[str] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, str, str]:
    """Load Step 15 outputs (YOY comparison and historical reference) in a period-aware way.

    Returns: (yoy_df, historical_ref_df, target_label, baseline_label)
    """
    # Determine labels
    baseline_yyyymm = baseline_yyyymm or _compute_baseline_yyyymm(target_yyyymm)
    baseline_period = baseline_period or target_period
    target_label = f"{target_yyyymm}{target_period}"
    baseline_label = f"{baseline_yyyymm}{baseline_period}"

    # Resolve file paths
    manifest = get_manifest().manifest if callable(get_manifest) else {}
    step15_outputs = (manifest.get("steps", {}).get("step15", {}).get("outputs", {}) if isinstance(manifest, dict) else {})

    # Prefer period-specific keys registered by Step 15 (yoy_comparison_{baseline}, historical_reference_{baseline})
    yoy_specific = step15_outputs.get(f"yoy_comparison_{baseline_label}", {}).get("file_path")
    hist_specific = step15_outputs.get(f"historical_reference_{baseline_label}", {}).get("file_path")
    yoy_generic = step15_outputs.get("yoy_comparison", {}).get("file_path")
    hist_generic = step15_outputs.get("historical_reference", {}).get("file_path")
    yoy_path = yoy_file_override or yoy_specific or yoy_generic
    hist_path = historical_file_override or hist_specific or hist_generic
    if yoy_specific:
        logger.info(f"Using period-specific YOY path from manifest: {yoy_specific}")
    elif yoy_generic and not yoy_file_override:
        logger.info(f"Using generic YOY path from manifest: {yoy_generic}")
    if hist_specific:
        logger.info(f"Using period-specific Historical path from manifest: {hist_specific}")
    elif hist_generic and not historical_file_override:
        logger.info(f"Using generic Historical path from manifest: {hist_generic}")

    # If manifest entries exist, ensure their baseline matches the requested baseline; otherwise ignore to avoid wrong-period data
    try:
        yoy_meta = {}
        if yoy_path:
            if yoy_specific and yoy_path == yoy_specific:
                yoy_meta = step15_outputs.get(f"yoy_comparison_{baseline_label}", {}).get("metadata", {})
            else:
                yoy_meta = step15_outputs.get("yoy_comparison", {}).get("metadata", {})
        if yoy_path and yoy_meta.get("baseline") and yoy_meta.get("baseline") != baseline_label:
            logger.info(
                f"Manifest YOY baseline {yoy_meta.get('baseline')} does not match requested {baseline_label}; ignoring manifest YOY entry"
            )
            yoy_path = None
    except Exception:
        pass
    try:
        hist_meta = {}
        if hist_path:
            if hist_specific and hist_path == hist_specific:
                hist_meta = step15_outputs.get(f"historical_reference_{baseline_label}", {}).get("metadata", {})
            else:
                hist_meta = step15_outputs.get("historical_reference", {}).get("metadata", {})
        if hist_path and hist_meta.get("baseline") and hist_meta.get("baseline") != baseline_label:
            logger.info(
                f"Manifest Historical baseline {hist_meta.get('baseline')} does not match requested {baseline_label}; ignoring manifest Historical entry"
            )
            hist_path = None
    except Exception:
        pass

    # Minimal fallback by pattern if manifest missing
    if not yoy_path or not os.path.exists(yoy_path):
        try:
            import glob
            candidates = sorted(glob.glob(f"output/year_over_year_comparison_{baseline_label}_*.csv"), key=os.path.getctime, reverse=True)
            yoy_path = candidates[0] if candidates else None
        except Exception:
            yoy_path = None
    if not hist_path or not os.path.exists(hist_path):
        try:
            import glob
            candidates = sorted(glob.glob(f"output/historical_reference_{baseline_label}_*.csv"), key=os.path.getctime, reverse=True)
            hist_path = candidates[0] if candidates else None
        except Exception:
            hist_path = None

    if not yoy_path or not os.path.exists(yoy_path):
        raise FileNotFoundError(f"Step 15 YOY comparison not found for {baseline_label}. Provide --yoy-file or ensure manifest is updated.")
    if not hist_path or not os.path.exists(hist_path):
        raise FileNotFoundError(f"Step 15 historical reference not found for {baseline_label}. Provide --historical-file or ensure manifest is updated.")

    logger.info(f"Loading YOY comparison: {yoy_path}")
    yoy_df = pd.read_csv(yoy_path)
    logger.info(f"Loaded {len(yoy_df):,} YOY rows")

    logger.info(f"Loading historical reference: {hist_path}")
    historical_ref_df = pd.read_csv(hist_path)
    logger.info(f"Loaded {len(historical_ref_df):,} historical reference rows")

    # Standardize key columns
    if 'Store_Group_Name' not in yoy_df.columns and 'store_group' in yoy_df.columns:
        yoy_df['Store_Group_Name'] = yoy_df['store_group']
    if 'Category' in historical_ref_df.columns and 'category' not in historical_ref_df.columns:
        historical_ref_df = historical_ref_df.rename(columns={'Category': 'category'})
    if 'Subcategory' in historical_ref_df.columns and 'sub_category' not in historical_ref_df.columns:
        historical_ref_df = historical_ref_df.rename(columns={'Subcategory': 'sub_category'})

    return yoy_df, historical_ref_df, target_label, baseline_label

def create_store_groups(df: pd.DataFrame) -> pd.DataFrame:
    """Create store groups using consistent logic."""
    df_with_groups = df.copy()
    # Disabled synthetic grouping in production; require upstream-provided group or mark unknown
    df_with_groups['store_group'] = df_with_groups.get('Store_Group_Name', pd.Series(index=df_with_groups.index, dtype=object))
    missing_mask = df_with_groups['store_group'].isna()
    if missing_mask.any():
        df_with_groups.loc[missing_mask, 'store_group'] = 'Store Group Unknown'
    return df_with_groups

def create_summary_comparison(historical_df: pd.DataFrame, current_df: pd.DataFrame) -> pd.DataFrame:
    """Create high-level summary comparison."""
    
    logger.info("Creating summary comparison...")
    
    # Historical summary
    hist_summary = {
        'Period': 'Historical',
        'Total_Records': len(historical_df),
        'Unique_Stores': historical_df['str_code'].nunique(),
        'Unique_SPUs': historical_df['spu_code'].nunique(),
        'Unique_Categories': historical_df['cate_name'].nunique(),
        'Unique_SubCategories': historical_df['sub_cate_name'].nunique(),
        'Total_Sales_Amount': pd.to_numeric(historical_df.get('spu_sales_amt'), errors='coerce').sum(min_count=1),
        'Total_Quantity': pd.to_numeric(historical_df.get('quantity'), errors='coerce').sum(min_count=1),
        'Avg_Sales_Per_SPU': (pd.to_numeric(historical_df.get('spu_sales_amt'), errors='coerce').sum(min_count=1) / historical_df['spu_code'].nunique()) if historical_df['spu_code'].nunique() else np.nan,
        'Avg_Quantity_Per_SPU': (pd.to_numeric(historical_df.get('quantity'), errors='coerce').sum(min_count=1) / historical_df['spu_code'].nunique()) if historical_df['spu_code'].nunique() else np.nan,
        'Avg_Unit_Price': (pd.to_numeric(historical_df.get('spu_sales_amt'), errors='coerce').sum(min_count=1) / pd.to_numeric(historical_df.get('quantity'), errors='coerce').sum(min_count=1)) if pd.to_numeric(historical_df.get('quantity'), errors='coerce').sum(min_count=1) else np.nan
    }
    
    # Current summary
    curr_summary = {
        'Period': 'Current',
        'Total_Records': len(current_df),
        'Unique_Stores': current_df['str_code'].nunique(),
        'Unique_SPUs': current_df['spu_code'].nunique(),
        'Unique_Categories': current_df['cate_name'].nunique(),
        'Unique_SubCategories': current_df['sub_cate_name'].nunique(),
        'Total_Sales_Amount': pd.to_numeric(current_df.get('spu_sales_amt'), errors='coerce').sum(min_count=1),
        'Total_Quantity': pd.to_numeric(current_df.get('quantity'), errors='coerce').sum(min_count=1),
        'Avg_Sales_Per_SPU': (pd.to_numeric(current_df.get('spu_sales_amt'), errors='coerce').sum(min_count=1) / current_df['spu_code'].nunique()) if current_df['spu_code'].nunique() else np.nan,
        'Avg_Quantity_Per_SPU': (pd.to_numeric(current_df.get('quantity'), errors='coerce').sum(min_count=1) / current_df['spu_code'].nunique()) if current_df['spu_code'].nunique() else np.nan,
        'Avg_Unit_Price': (pd.to_numeric(current_df.get('spu_sales_amt'), errors='coerce').sum(min_count=1) / pd.to_numeric(current_df.get('quantity'), errors='coerce').sum(min_count=1)) if pd.to_numeric(current_df.get('quantity'), errors='coerce').sum(min_count=1) else np.nan
    }
    
    # Create comparison DataFrame
    comparison_df = pd.DataFrame([hist_summary, curr_summary])
    
    # Calculate changes
    changes = {}
    for col in ['Total_Records', 'Unique_Stores', 'Unique_SPUs', 'Unique_Categories', 
                'Unique_SubCategories', 'Total_Sales_Amount', 'Total_Quantity', 
                'Avg_Sales_Per_SPU', 'Avg_Quantity_Per_SPU', 'Avg_Unit_Price']:
        hist_val = comparison_df.iloc[0][col]
        curr_val = comparison_df.iloc[1][col]
        change = (curr_val - hist_val) if pd.notna(hist_val) and pd.notna(curr_val) else np.nan
        change_pct = (change / hist_val * 100) if (pd.notna(hist_val) and hist_val != 0) else np.nan
        changes[col] = change
        changes[f'{col}_Pct'] = change_pct
    
    # Add change row
    changes['Period'] = 'Change (Current - Historical)'
    comparison_df = pd.concat([comparison_df, pd.DataFrame([changes])], ignore_index=True)
    
    return comparison_df

def create_category_comparison(historical_df: pd.DataFrame, current_df: pd.DataFrame) -> pd.DataFrame:
    """Create category-level comparison."""
    
    logger.info("Creating category comparison...")
    
    # Historical category analysis
    hist_cat = historical_df.groupby('cate_name').agg({
        'spu_code': 'nunique',
        'spu_sales_amt': 'sum',
        'quantity': 'sum',
        'str_code': 'nunique'
    }).reset_index()
    hist_cat.columns = ['Category', 'Historical_SPU_Count', 'Historical_Sales', 'Historical_Quantity', 'Historical_Stores']
    hist_cat['Historical_Avg_Sales_Per_SPU'] = np.where(hist_cat['Historical_SPU_Count'] > 0, hist_cat['Historical_Sales'] / hist_cat['Historical_SPU_Count'], np.nan)
    
    # Current category analysis
    curr_cat = current_df.groupby('cate_name').agg({
        'spu_code': 'nunique',
        'spu_sales_amt': 'sum',
        'quantity': 'sum',
        'str_code': 'nunique'
    }).reset_index()
    curr_cat.columns = ['Category', 'Current_SPU_Count', 'Current_Sales', 'Current_Quantity', 'Current_Stores']
    curr_cat['Current_Avg_Sales_Per_SPU'] = np.where(curr_cat['Current_SPU_Count'] > 0, curr_cat['Current_Sales'] / curr_cat['Current_SPU_Count'], np.nan)
    
    # Merge and calculate changes
    category_comparison = hist_cat.merge(curr_cat, on='Category', how='outer')
    
    category_comparison['SPU_Count_Change'] = category_comparison['Current_SPU_Count'] - category_comparison['Historical_SPU_Count']
    category_comparison['SPU_Count_Change_Pct'] = np.where(category_comparison['Historical_SPU_Count'] > 0, (category_comparison['SPU_Count_Change'] / category_comparison['Historical_SPU_Count']) * 100, np.nan).round(1)
    
    category_comparison['Sales_Change'] = category_comparison['Current_Sales'] - category_comparison['Historical_Sales']
    category_comparison['Sales_Change_Pct'] = np.where(category_comparison['Historical_Sales'] > 0, (category_comparison['Sales_Change'] / category_comparison['Historical_Sales']) * 100, np.nan).round(1)
    
    category_comparison['Avg_Sales_Per_SPU_Change'] = category_comparison['Current_Avg_Sales_Per_SPU'] - category_comparison['Historical_Avg_Sales_Per_SPU']
    category_comparison['Avg_Sales_Per_SPU_Change_Pct'] = np.where(category_comparison['Historical_Avg_Sales_Per_SPU'] > 0, (category_comparison['Avg_Sales_Per_SPU_Change'] / category_comparison['Historical_Avg_Sales_Per_SPU']) * 100, np.nan).round(1)
    
    # Sort by historical sales
    category_comparison = category_comparison.sort_values('Historical_Sales', ascending=False)
    
    return category_comparison

def create_store_group_comparison(historical_df: pd.DataFrame, current_df: pd.DataFrame) -> pd.DataFrame:
    """Create store group comparison."""
    
    logger.info("Creating store group comparison...")
    
    # Add store groups
    historical_grouped = create_store_groups(historical_df)
    current_grouped = create_store_groups(current_df)
    
    # Historical store group analysis
    hist_sg = historical_grouped.groupby('store_group').agg({
        'spu_code': 'nunique',
        'spu_sales_amt': 'sum',
        'quantity': 'sum',
        'str_code': 'nunique'
    }).reset_index()
    hist_sg.columns = ['Store_Group', 'Historical_SPU_Count', 'Historical_Sales', 'Historical_Quantity', 'Historical_Stores']
    hist_sg['Historical_Avg_Sales_Per_SPU'] = np.where(hist_sg['Historical_SPU_Count'] > 0, hist_sg['Historical_Sales'] / hist_sg['Historical_SPU_Count'], np.nan)
    
    # Current store group analysis
    curr_sg = current_grouped.groupby('store_group').agg({
        'spu_code': 'nunique',
        'spu_sales_amt': 'sum',
        'quantity': 'sum',
        'str_code': 'nunique'
    }).reset_index()
    curr_sg.columns = ['Store_Group', 'Current_SPU_Count', 'Current_Sales', 'Current_Quantity', 'Current_Stores']
    curr_sg['Current_Avg_Sales_Per_SPU'] = np.where(curr_sg['Current_SPU_Count'] > 0, curr_sg['Current_Sales'] / curr_sg['Current_SPU_Count'], np.nan)
    
    # Merge and calculate changes
    sg_comparison = hist_sg.merge(curr_sg, on='Store_Group', how='outer')
    
    sg_comparison['SPU_Count_Change'] = sg_comparison['Current_SPU_Count'] - sg_comparison['Historical_SPU_Count']
    sg_comparison['SPU_Count_Change_Pct'] = np.where(sg_comparison['Historical_SPU_Count'] > 0, (sg_comparison['SPU_Count_Change'] / sg_comparison['Historical_SPU_Count']) * 100, np.nan).round(1)
    
    sg_comparison['Sales_Change'] = sg_comparison['Current_Sales'] - sg_comparison['Historical_Sales']
    sg_comparison['Sales_Change_Pct'] = np.where(sg_comparison['Historical_Sales'] > 0, (sg_comparison['Sales_Change'] / sg_comparison['Historical_Sales']) * 100, np.nan).round(1)
    
    # Sort by store group
    sg_comparison = sg_comparison.sort_values('Store_Group')
    
    return sg_comparison

def create_summary_period_aware(yoy_df: pd.DataFrame, baseline_label: str, target_label: str) -> pd.DataFrame:
    """Create high-level summary using Step 15 YOY data (period-aware)."""
    logger.info("Creating period-aware summary from YOY data...")
    # Fill numeric NaNs
    cols_numeric = [
        'historical_spu_count','historical_total_sales','historical_total_quantity','historical_store_count',
        'Current_SPU_Quantity','Total_Current_Sales'
    ]
    for c in cols_numeric:
        if c in yoy_df.columns:
            yoy_df[c] = yoy_df[c].fillna(0)

    # Historical aggregates
    hist_spus = yoy_df.get('historical_spu_count', pd.Series(dtype=float)).sum()
    hist_sales = yoy_df.get('historical_total_sales', pd.Series(dtype=float)).sum()
    hist_qty = yoy_df.get('historical_total_quantity', pd.Series(dtype=float)).sum()
    # Current aggregates
    curr_spus = yoy_df.get('Current_SPU_Quantity', pd.Series(dtype=float)).sum()
    curr_sales = yoy_df.get('Total_Current_Sales', pd.Series(dtype=float)).sum()

    unique_groups = (yoy_df['Store_Group_Name'].nunique() if 'Store_Group_Name' in yoy_df.columns
                     else yoy_df.get('store_group', pd.Series(dtype=object)).nunique())
    unique_categories = yoy_df.get('category', pd.Series(dtype=object)).nunique()
    unique_subcategories = yoy_df.get('sub_category', pd.Series(dtype=object)).nunique()

    hist_avg_sales_per_spu = (hist_sales / hist_spus) if hist_spus else 0
    curr_avg_sales_per_spu = (curr_sales / curr_spus) if curr_spus else 0

    rows = [
        {
            'Period': f'{baseline_label} (Historical)',
            'Total_Combinations': len(yoy_df),
            'Unique_Store_Groups': unique_groups,
            'Unique_Categories': unique_categories,
            'Unique_SubCategories': unique_subcategories,
            'Total_Sales_Historical': hist_sales,
            'Total_SPUs_Historical': hist_spus,
            'Avg_Sales_Per_SPU_Historical': hist_avg_sales_per_spu,
            'Total_Sales_Current': np.nan,
            'Total_SPUs_Current': np.nan,
            'Avg_Sales_Per_SPU_Current': np.nan,
        },
        {
            'Period': f'{target_label} (Current)',
            'Total_Combinations': len(yoy_df),
            'Unique_Store_Groups': unique_groups,
            'Unique_Categories': unique_categories,
            'Unique_SubCategories': unique_subcategories,
            'Total_Sales_Historical': np.nan,
            'Total_SPUs_Historical': np.nan,
            'Avg_Sales_Per_SPU_Historical': np.nan,
            'Total_Sales_Current': curr_sales,
            'Total_SPUs_Current': curr_spus,
            'Avg_Sales_Per_SPU_Current': curr_avg_sales_per_spu,
        }
    ]

    summary_df = pd.DataFrame(rows)

    # Change row (where both exist)
    change = {
        'Period': f'Change ({target_label} - {baseline_label})',
        'Total_Combinations': 0,
        'Unique_Store_Groups': 0,
        'Unique_Categories': 0,
        'Unique_SubCategories': 0,
        'Total_Sales_Historical': 0,
        'Total_SPUs_Historical': 0,
        'Avg_Sales_Per_SPU_Historical': 0,
        'Total_Sales_Current': curr_sales - hist_sales,
        'Total_SPUs_Current': curr_spus - hist_spus,
        'Avg_Sales_Per_SPU_Current': curr_avg_sales_per_spu - hist_avg_sales_per_spu,
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([change])], ignore_index=True)
    return summary_df

def create_category_comparison_from_yoy(yoy_df: pd.DataFrame) -> pd.DataFrame:
    """Category-level comparison from YOY data."""
    logger.info("Creating category comparison from YOY data...")
    df = yoy_df.copy()
    for c in ['historical_spu_count','historical_total_sales','Current_SPU_Quantity','Total_Current_Sales']:
        if c in df.columns:
            df[c] = df[c].fillna(0)
    cat = df.groupby('category').agg({
        'historical_spu_count': 'sum',
        'historical_total_sales': 'sum',
        'Current_SPU_Quantity': 'sum',
        'Total_Current_Sales': 'sum'
    }).reset_index()
    cat['SPU_Count_Change'] = cat['Current_SPU_Quantity'] - cat['historical_spu_count']
    cat['SPU_Count_Change_Pct'] = (cat['SPU_Count_Change'] / cat['historical_spu_count'] * 100).replace([np.inf, -np.inf], np.nan).round(1)
    cat['Sales_Change'] = cat['Total_Current_Sales'] - cat['historical_total_sales']
    cat['Sales_Change_Pct'] = (cat['Sales_Change'] / cat['historical_total_sales'] * 100).replace([np.inf, -np.inf], np.nan).round(1)
    cat = cat.sort_values('historical_total_sales', ascending=False)
    return cat

def create_store_group_comparison_from_yoy(yoy_df: pd.DataFrame) -> pd.DataFrame:
    """Store group comparison from YOY data."""
    logger.info("Creating store group comparison from YOY data...")
    df = yoy_df.copy()
    for c in ['historical_spu_count','historical_total_sales','Current_SPU_Quantity','Total_Current_Sales']:
        if c in df.columns:
            df[c] = df[c].fillna(0)
    group_col = 'Store_Group_Name' if 'Store_Group_Name' in df.columns else 'store_group'
    # Diagnostics: mapping coverage
    try:
        total_rows = int(len(df))
        unknown_na = int(df[group_col].isna().sum()) if group_col in df.columns else 0
        unknown_named = int((df[group_col] == 'Store Group Unknown').sum()) if group_col in df.columns else 0
        logger.info(
            f"YOY store-group coverage: {total_rows - (unknown_na + unknown_named)}/{total_rows} mapped, "
            f"{unknown_na + unknown_named} Unknown/NA"
        )
    except Exception:
        pass
    sg = df.groupby(group_col).agg({
        'historical_spu_count': 'sum',
        'historical_total_sales': 'sum',
        'Current_SPU_Quantity': 'sum',
        'Total_Current_Sales': 'sum'
    }).reset_index()
    sg = sg.rename(columns={group_col: 'Store_Group'})
    sg['SPU_Count_Change'] = sg['Current_SPU_Quantity'] - sg['historical_spu_count']
    sg['Sales_Change'] = sg['Total_Current_Sales'] - sg['historical_total_sales']
    with np.errstate(divide='ignore', invalid='ignore'):
        sg['Sales_Change_Pct'] = np.where(sg['historical_total_sales'] > 0, (sg['Sales_Change'] / sg['historical_total_sales']) * 100, np.nan).round(1)
    sg = sg.sort_values('Store_Group')
    return sg

def _try_load_raw_spu_csv(path: Optional[str]) -> Optional[pd.DataFrame]:
    """Try to load a raw SPU CSV if path exists; return None otherwise."""
    if not path:
        return None
    try:
        if os.path.exists(path):
            df = pd.read_csv(path, dtype={'str_code': str, 'spu_code': str})
            # normalize expected columns
            for col in ['spu_sales_amt', 'quantity']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            # standardize category columns if needed
            if {'Category', 'Subcategory'}.issubset(df.columns):
                df = df.rename(columns={'Category': 'cate_name', 'Subcategory': 'sub_cate_name'})
            return df
    except Exception as e:
        logger.warning(f"Failed to load raw SPU CSV at {path}: {e}")
    return None

def load_optional_raw_spu_data(
    baseline_label: str,
    target_yyyymm: str,
    target_period: str,
    historical_raw_override: Optional[str] = None,
    current_raw_override: Optional[str] = None,
) -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame]]:
    """Attempt to load baseline and current raw SPU sales CSVs using period-aware config; fallback to direct period files."""
    # Historical (baseline)
    hist = _try_load_raw_spu_csv(historical_raw_override)
    if hist is None:
        try:
            b_yyyymm = str(baseline_label)[:-1]
            b_period = str(baseline_label)[-1]
            api_hist = get_api_data_files(b_yyyymm, b_period)
            candidates = [api_hist.get('spu_sales'), api_hist.get('complete_spu_sales'), f"data/api_data/complete_spu_sales_{baseline_label}.csv"]
            chosen = next((p for p in candidates if p and os.path.exists(p)), None)
            hist = _try_load_raw_spu_csv(chosen)
        except Exception:
            hist = _try_load_raw_spu_csv(f"data/api_data/complete_spu_sales_{baseline_label}.csv")
    # Current (target)
    curr = _try_load_raw_spu_csv(current_raw_override)
    if curr is None:
        try:
            api_curr = get_api_data_files(target_yyyymm, target_period)
            candidates = [api_curr.get('spu_sales'), api_curr.get('complete_spu_sales'), f"data/api_data/complete_spu_sales_{target_yyyymm}{target_period}.csv"]
            chosen = next((p for p in candidates if p and os.path.exists(p)), None)
            curr = _try_load_raw_spu_csv(chosen)
        except Exception:
            curr = _try_load_raw_spu_csv(f"data/api_data/complete_spu_sales_{target_yyyymm}{target_period}.csv")
    if hist is None:
        logger.info("Baseline raw SPU data not available; Top SPUs (Historical) will be skipped.")
    if curr is None:
        logger.info("Current raw SPU data not available; Top SPUs (Current) will be skipped.")
    return hist, curr

def create_top_performers_analysis(historical_df: pd.DataFrame, current_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Create top performers analysis."""
    
    logger.info("Creating top performers analysis...")
    
    results = {}
    
    # Top SPUs by sales (historical)
    hist_top_spus = historical_df.groupby('spu_code').agg({
        'spu_sales_amt': 'sum',
        'quantity': 'sum',
        'str_code': 'nunique'
    }).reset_index()
    hist_top_spus['avg_sales_per_store'] = hist_top_spus['spu_sales_amt'] / hist_top_spus['str_code']
    hist_top_spus = hist_top_spus.sort_values('spu_sales_amt', ascending=False).head(20)
    hist_top_spus['Period'] = 'May 2025_Historical'
    results['top_spus_historical'] = hist_top_spus
    
    # Top SPUs by sales (current)
    curr_top_spus = current_df.groupby('spu_code').agg({
        'spu_sales_amt': 'sum',
        'quantity': 'sum',
        'str_code': 'nunique'
    }).reset_index()
    curr_top_spus['avg_sales_per_store'] = curr_top_spus['spu_sales_amt'] / curr_top_spus['str_code']
    curr_top_spus = curr_top_spus.sort_values('spu_sales_amt', ascending=False).head(20)
    curr_top_spus['Period'] = 'July 2025_Current'
    results['top_spus_current'] = curr_top_spus
    
    # Top categories by growth
    hist_cat_totals = historical_df.groupby('cate_name')['spu_sales_amt'].sum()
    curr_cat_totals = current_df.groupby('cate_name')['spu_sales_amt'].sum()
    
    category_growth = pd.DataFrame({
        'Category': hist_cat_totals.index,
        'Historical_Sales': hist_cat_totals.values,
        'Current_Sales': curr_cat_totals.reindex(hist_cat_totals.index, fill_value=0).values
    })
    category_growth['Growth'] = category_growth['Current_Sales'] - category_growth['Historical_Sales']
    category_growth['Growth_Pct'] = (category_growth['Growth'] / category_growth['Historical_Sales'] * 100).round(1)
    category_growth = category_growth.sort_values('Growth_Pct', ascending=False)
    results['category_growth'] = category_growth
    
    return results

def create_top_performers_analysis_from_raw(
    historical_raw_df: Optional[pd.DataFrame],
    current_raw_df: Optional[pd.DataFrame],
    baseline_label: str,
    target_label: str,
) -> Dict[str, pd.DataFrame]:
    """Create Top SPUs and Category Growth from raw SPU-level data if available."""
    results: Dict[str, pd.DataFrame] = {}
    if historical_raw_df is not None:
        try:
            hist_top_spus = historical_raw_df.groupby('spu_code').agg({
                'spu_sales_amt': 'sum',
                'quantity': 'sum',
                'str_code': 'nunique'
            }).reset_index()
            hist_top_spus['avg_sales_per_store'] = hist_top_spus['spu_sales_amt'] / hist_top_spus['str_code']
            hist_top_spus = hist_top_spus.sort_values('spu_sales_amt', ascending=False).head(50)
            hist_top_spus['Period'] = f'{baseline_label}_Historical'
            results['Top_SPUs_Historical'] = hist_top_spus
        except Exception as e:
            logger.warning(f"Top SPUs (Historical) failed: {e}")
    if current_raw_df is not None:
        try:
            curr_top_spus = current_raw_df.groupby('spu_code').agg({
                'spu_sales_amt': 'sum',
                'quantity': 'sum',
                'str_code': 'nunique'
            }).reset_index()
            curr_top_spus['avg_sales_per_store'] = curr_top_spus['spu_sales_amt'] / curr_top_spus['str_code']
            curr_top_spus = curr_top_spus.sort_values('spu_sales_amt', ascending=False).head(50)
            curr_top_spus['Period'] = f'{target_label}_Current'
            results['Top_SPUs_Current'] = curr_top_spus
        except Exception as e:
            logger.warning(f"Top SPUs (Current) failed: {e}")
    # Category growth from raw (if both sides available)
    if historical_raw_df is not None and current_raw_df is not None:
        try:
            hist_cat_totals = historical_raw_df.groupby('cate_name')['spu_sales_amt'].sum()
            curr_cat_totals = current_raw_df.groupby('cate_name')['spu_sales_amt'].sum()
            category_growth = pd.DataFrame({
                'Category': hist_cat_totals.index,
                'Historical_Sales': hist_cat_totals.values,
                'Current_Sales': curr_cat_totals.reindex(hist_cat_totals.index, fill_value=0).values
            })
            category_growth['Growth'] = category_growth['Current_Sales'] - category_growth['Historical_Sales']
            with np.errstate(divide='ignore', invalid='ignore'):
                category_growth['Growth_Pct'] = np.where(category_growth['Historical_Sales'] > 0, (category_growth['Growth'] / category_growth['Historical_Sales']) * 100, np.nan).round(1)
            category_growth = category_growth.sort_values('Growth_Pct', ascending=False)
            results['Category_Growth_Raw'] = category_growth
        except Exception as e:
            logger.warning(f"Category Growth (Raw) failed: {e}")
    return results

def build_focus_view_for_keywords(yoy_df: pd.DataFrame, keywords: List[str]) -> pd.DataFrame:
    """Create a filtered view of YOY rows matching any keyword in category/sub-category (case-insensitive)."""
    df = yoy_df.copy()
    df['category'] = df.get('category', pd.Series(index=df.index, dtype=object)).astype(str)
    df['sub_category'] = df.get('sub_category', pd.Series(index=df.index, dtype=object)).astype(str)
    mask = False
    for kw in keywords:
        kw_lower = str(kw).lower()
        mask = mask | df['category'].str.lower().str.contains(kw_lower, na=False) | df['sub_category'].str.lower().str.contains(kw_lower, na=False)
    filtered = df[mask].copy()
    return filtered

def classify_fashion_vs_basic(
    yoy_df: pd.DataFrame,
    fashion_keywords: List[str],
    basic_keywords: List[str],
    hybrid_threshold: float = 0.15,
) -> pd.DataFrame:
    """Classify each YOY row with Fashion/Basic hit counts and ratio; mark Hybrid if fairly split.

    - Fashion_Hits / Basic_Hits: count of unique keyword matches across category/sub_category
    - Fashion_Share_Pct: 100 * Fashion_Hits / (Fashion_Hits + Basic_Hits)
    - Segment: Fashion | Basic | Hybrid | Unclassified
    """
    df = yoy_df.copy()
    df['category'] = df.get('category', pd.Series(index=df.index, dtype=object)).astype(str)
    df['sub_category'] = df.get('sub_category', pd.Series(index=df.index, dtype=object)).astype(str)

    f_hits = pd.Series(0, index=df.index, dtype=int)
    b_hits = pd.Series(0, index=df.index, dtype=int)

    # Deduplicate keywords, ignore empties
    f_kw = [k.lower() for k in fashion_keywords if str(k).strip()]
    b_kw = [k.lower() for k in basic_keywords if str(k).strip()]
    f_kw = list(dict.fromkeys(f_kw))
    b_kw = list(dict.fromkeys(b_kw))

    for kw in f_kw:
        match = df['category'].str.lower().str.contains(kw, na=False) | df['sub_category'].str.lower().str.contains(kw, na=False)
        f_hits = f_hits.add(match.astype(int), fill_value=0).astype(int)
    for kw in b_kw:
        match = df['category'].str.lower().str.contains(kw, na=False) | df['sub_category'].str.lower().str.contains(kw, na=False)
        b_hits = b_hits.add(match.astype(int), fill_value=0).astype(int)

    total_hits = f_hits + b_hits
    with np.errstate(divide='ignore', invalid='ignore'):
        fashion_share = np.where(total_hits > 0, f_hits / total_hits, np.nan)

    # Segment classification
    seg = np.where(total_hits == 0, 'Unclassified',
          np.where(np.abs(fashion_share - 0.5) <= hybrid_threshold, 'Hybrid',
          np.where(fashion_share > 0.5, 'Fashion', 'Basic')))

    out = df.copy()
    out['Fashion_Hits'] = f_hits
    out['Basic_Hits'] = b_hits
    out['Fashion_Share_Pct'] = (fashion_share * 100).round(1)
    out['Segment'] = seg
    return out

def build_segment_summary(segment_df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate YOY metrics by Segment similar to store group/category summaries."""
    df = segment_df.copy()
    for c in ['historical_spu_count','historical_total_sales','Current_SPU_Quantity','Total_Current_Sales']:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0)
    seg = df.groupby('Segment').agg({
        'historical_spu_count': 'sum',
        'historical_total_sales': 'sum',
        'Current_SPU_Quantity': 'sum',
        'Total_Current_Sales': 'sum',
    }).reset_index()
    seg['SPU_Count_Change'] = seg['Current_SPU_Quantity'] - seg['historical_spu_count']
    seg['Sales_Change'] = seg['Total_Current_Sales'] - seg['historical_total_sales']
    with np.errstate(divide='ignore', invalid='ignore'):
        seg['Sales_Change_Pct'] = np.where(seg['historical_total_sales'] > 0, (seg['Sales_Change'] / seg['historical_total_sales']) * 100, np.nan).round(1)
    seg = seg.sort_values('Segment')
    return seg

def _polish_excel_sheet(ws):
    """Apply common polish to a worksheet: freeze panes, header style, autofilter, column widths, number formats, conditional formatting."""
    if ws.max_row >= 2:
        ws.freeze_panes = 'A2'
    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    # AutoFilter
    ws.auto_filter.ref = ws.dimensions
    # Auto width (cap width)
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 200) + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            val_str = str(val) if val is not None else ""
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))
    # Number formats and conditional formatting on change columns
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
    for name, col_idx in header_map.items():
        name_str = str(name) if name is not None else ""
        name_lower = name_str.lower()
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{ws.max_row}"
        if any(token in name_lower for token in ["sales", "amount", "revenue"]):
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in r:
                    cell.number_format = '#,##0'
        if any(token in name_lower for token in ["quantity", "count", "stores"]):
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in r:
                    cell.number_format = '#,##0'
        if name_lower.endswith('_pct') or 'pct' in name_lower or name_str.endswith('%'):
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                for cell in r:
                    cell.number_format = '0.0'
        if 'change' in name_lower:
            try:
                ws.conditional_formatting.add(data_range, CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')))
                ws.conditional_formatting.add(data_range, CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')))
            except Exception:
                pass


def save_excel_analysis(
    summary_df: pd.DataFrame,
    category_df: pd.DataFrame,
    store_group_df: pd.DataFrame,
    yoy_df: pd.DataFrame,
    historical_ref_df: pd.DataFrame,
    target_label: str,
    extra_sheets: Optional[Dict[str, pd.DataFrame]] = None,
) -> str:
    """Save comprehensive Excel analysis (period-aware) to output/ and return file path."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs("output", exist_ok=True)
    
    # DUAL OUTPUT PATTERN
    timestamped_excel_file = f"output/spreadsheet_comparison_analysis_{target_label}_{timestamp}.xlsx"
    generic_excel_file = f"output/spreadsheet_comparison_analysis_{target_label}.xlsx"

    logger.info(f"Saving Excel analysis to: {timestamped_excel_file}")

    # Save timestamped version (for backup/inspection)
    with pd.ExcelWriter(timestamped_excel_file, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        category_df.to_excel(writer, sheet_name='Category_Comparison', index=False)
        store_group_df.to_excel(writer, sheet_name='Store_Group_Comparison', index=False)
        yoy_df.to_excel(writer, sheet_name='YOY_Comparison_Raw', index=False)
        historical_ref_df.to_excel(writer, sheet_name='Historical_Reference_Raw', index=False)
        # Extra sheets
        if extra_sheets:
            for sheet_name, df in extra_sheets.items():
                try:
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                except Exception as e:
                    logger.warning(f"Could not write extra sheet '{sheet_name}': {e}")
        # Polish all sheets
        try:
            for ws in writer.book.worksheets:
                _polish_excel_sheet(ws)
        except Exception as e:
            logger.warning(f"Excel polish step encountered an issue: {e}")

    logger.info(f"Timestamped Excel analysis saved: {timestamped_excel_file}")
    
    # Create symlink for generic version (for pipeline flow)
    if os.path.exists(generic_excel_file) or os.path.islink(generic_excel_file):
        os.remove(generic_excel_file)
    os.symlink(os.path.basename(timestamped_excel_file), generic_excel_file)
    
    logger.info(f"Generic Excel symlink created: {generic_excel_file} -> {timestamped_excel_file}")
    
    excel_file = timestamped_excel_file
    return excel_file

def main():
    """Main execution function."""
    
    logger.info("Starting Step 16: Spreadsheet Comparison Analysis (period-aware)...")

    try:
        # Parse args
        args = _parse_args()
        target_yyyymm = args.target_yyyymm
        target_period = args.target_period
        baseline_yyyymm = args.baseline_yyyymm
        baseline_period = args.baseline_period

        # Load Step 15 outputs (YOY + historical reference)
        yoy_df, historical_ref_df, target_label, baseline_label = load_data(
            target_yyyymm,
            target_period,
            baseline_yyyymm,
            baseline_period,
            args.yoy_file,
            args.historical_file,
        )

        # Build comparisons
        summary_df = create_summary_period_aware(yoy_df, baseline_label, target_label)
        category_df = create_category_comparison_from_yoy(yoy_df)
        store_group_df = create_store_group_comparison_from_yoy(yoy_df)

        # Extra sheets container
        extra_sheets: Dict[str, pd.DataFrame] = {}
        info_messages: List[str] = []

        # Top by growth (from YOY aggregates)
        try:
            top_cats = category_df.sort_values('Sales_Change_Pct', ascending=False).head(50)
            extra_sheets['Top_Categories_Growth'] = top_cats
        except Exception:
            pass
        try:
            sg_growth = store_group_df.copy()
            if 'Sales_Change_Pct' in sg_growth.columns:
                sg_growth = sg_growth.sort_values('Sales_Change_Pct', ascending=False)
            else:
                sg_growth = sg_growth.sort_values('Sales_Change', ascending=False)
            extra_sheets['Top_Store_Groups_Growth'] = sg_growth.head(50)
        except Exception:
            pass

        # Optional: load raw SPU for Top SPUs sheets
        hist_raw_df, curr_raw_df = load_optional_raw_spu_data(
            baseline_label,
            target_yyyymm,
            target_period,
            args.historical_raw_file,
            args.current_raw_file,
        )
        # Diagnostics for missing raw inputs
        if hist_raw_df is None:
            looked_for_hist = args.historical_raw_file or f"data/api_data/complete_spu_sales_{baseline_label}.csv"
            info_messages.append(
                f"Top_SPUs_Historical skipped: baseline raw SPU file not found or unreadable. Looked for: {looked_for_hist}"
            )
        if curr_raw_df is None:
            looked_for_curr = args.current_raw_file or f"data/api_data/complete_spu_sales_{target_yyyymm}{target_period}.csv"
            info_messages.append(
                f"Top_SPUs_Current / Category_Growth_Raw skipped: current raw SPU file not found or unreadable. Looked for: {looked_for_curr}"
            )
        raw_top = create_top_performers_analysis_from_raw(hist_raw_df, curr_raw_df, baseline_label, target_label)
        for k, v in raw_top.items():
            extra_sheets[k] = v

        # Fashion vs Basic segmentation (no makeup)
        try:
            # Defaults
            default_fashion_keywords = [
                # English
                'fashion',
                # Chinese common apparel terms
                '时尚','服饰','T恤','POLO','衬衫','牛仔','休闲','卫衣','毛衣','针织','风衣','大衣','夹克','茄克','外套','裤','裙','连衣裙','配饰','鞋','箱包'
            ]
            default_basic_keywords = [
                # English
                'basic','basics',
                # Chinese basics
                '基础','基础款','基本款','打底','百搭','素色','纯色','经典'
            ]
            user_fashion = [k.strip() for k in str(args.fashion_keywords).split(',')] if getattr(args, 'fashion_keywords', None) else []
            user_basic = [k.strip() for k in str(args.basic_keywords).split(',')] if getattr(args, 'basic_keywords', None) else []
            fashion_keywords = list(dict.fromkeys(user_fashion + default_fashion_keywords))
            basic_keywords = list(dict.fromkeys(user_basic + default_basic_keywords))

            # Optional deprecation warning if legacy flag provided
            try:
                if getattr(args, 'focus_keywords', None):
                    logger.warning("--focus-keywords is deprecated and ignored. Use --fashion-keywords / --basic-keywords.")
                    info_messages.append("--focus-keywords is deprecated and ignored. Use --fashion-keywords / --basic-keywords.")
            except Exception:
                pass

            hybrid_th = getattr(args, 'hybrid_threshold', 0.15) if hasattr(args, 'hybrid_threshold') else 0.15
            classified = classify_fashion_vs_basic(yoy_df, fashion_keywords, basic_keywords, hybrid_threshold=hybrid_th)
            seg_summary = build_segment_summary(classified[classified['Segment'].isin(['Fashion','Basic','Hybrid'])])
            extra_sheets['Fashion_vs_Basic'] = seg_summary

            # Include Hybrid in both focus sheets to aid review
            fashion_focus = classified[classified['Segment'].isin(['Fashion','Hybrid'])].copy()
            basic_focus = classified[classified['Segment'].isin(['Basic','Hybrid'])].copy()
            extra_sheets['Fashion_Focus'] = fashion_focus
            extra_sheets['Basic_Focus'] = basic_focus

            if len(fashion_focus) == 0:
                info_messages.append("Fashion_Focus has 0 rows. Adjust --fashion-keywords if needed.")
            if len(basic_focus) == 0:
                info_messages.append("Basic_Focus has 0 rows. Adjust --basic-keywords if needed.")
            try:
                counts = classified['Segment'].value_counts(dropna=False).to_dict()
                info_messages.append(
                    f"Fashion vs Basic segmentation: Hybrid threshold=±{int(hybrid_th*100)}% around 50%. Segment counts: {counts}"
                )
            except Exception:
                pass
        except Exception as e:
            info_messages.append(f"Fashion vs Basic segmentation failed: {e}")

        # Attach diagnostics/info sheet if any
        if info_messages:
            try:
                extra_sheets['Info'] = pd.DataFrame({'Message': info_messages})
            except Exception:
                pass

        # Save Excel analysis
        excel_file = save_excel_analysis(
            summary_df,
            category_df,
            store_group_df,
            yoy_df,
            historical_ref_df,
            target_label,
            extra_sheets=extra_sheets,
        )

        # Register output in manifest
        try:
            out_metadata = {
                # Period-aware identifiers
                "period_label": str(target_label),
                "target_yyyymm": str(target_yyyymm),
                "target_year": int(target_yyyymm[:4]),
                "target_month": int(target_yyyymm[4:6]),
                "target_period": target_period,
                # Baseline identifiers
                "baseline_label": str(baseline_label),
                "baseline_yyyymm": str(baseline_label)[:-1],
                "baseline_period": str(baseline_label)[-1],
                # Backward compatibility
                "baseline": baseline_label,
                # File info
                "file_format": "xlsx",
                "sheets": {
                    "base": [
                        "Summary",
                        "Category_Comparison",
                        "Store_Group_Comparison",
                        "YOY_Comparison_Raw",
                        "Historical_Reference_Raw",
                    ],
                    "extras": sorted(list(extra_sheets.keys())) if extra_sheets else [],
                },
                # Row counts per primary table
                "rows": {
                    "summary": int(len(summary_df)),
                    "category": int(len(category_df)),
                    "store_group": int(len(store_group_df)),
                    "yoy": int(len(yoy_df)),
                    "historical": int(len(historical_ref_df)),
                },
            }
            # Generic key (kept for backward compatibility; will reflect the most recent run)
            register_step_output(
                "step16",
                "comparison_workbook",
                excel_file,
                metadata=out_metadata,
            )
            # Period-specific key to avoid overwriting and enable downstream disambiguation
            register_step_output(
                "step16",
                f"comparison_workbook_{target_label}",
                excel_file,
                metadata=out_metadata,
            )
        except Exception as e:
            logger.warning(f"Could not register Step 16 output in manifest: {e}")

        # Print summary
        print("\n=== STEP 16: SPREADSHEET COMPARISON ANALYSIS (PERIOD-AWARE) ===")
        print(f"Baseline period: {baseline_label}")
        print(f"Current period:  {target_label}")
        print("")
        print("Summary metrics (overall):")
        try:
            hist_sales = summary_df.loc[summary_df['Period'].str.contains('Historical'), 'Total_Sales_Historical'].iloc[0]
            curr_sales = summary_df.loc[summary_df['Period'].str.contains('Current'), 'Total_Sales_Current'].iloc[0]
            print(f"  Historical Sales: ¥{hist_sales:,.0f}")
            print(f"  Current Sales:    ¥{curr_sales:,.0f}")
            print(f"  Sales Change:     ¥{(curr_sales - hist_sales):+,.0f}")
        except Exception:
            pass
        print("")
        print(f"Excel file created: {excel_file}")

        logger.info("Step 16: Spreadsheet Comparison Analysis completed successfully!")
        return excel_file

    except Exception as e:
        logger.error(f"Error in Step 16: {e}")
        raise

if __name__ == "__main__":
    main()
