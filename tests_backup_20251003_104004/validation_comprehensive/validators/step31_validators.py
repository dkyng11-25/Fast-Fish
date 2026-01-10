#!/usr/bin/env python3
"""
Step 31 Gap Analysis Workbook Validators

This module contains validators for gap analysis workbook validation from Step 31.

Author: Data Pipeline
Date: 2025-01-16
"""

import logging
from pathlib import Path
from typing import Dict, Any, List
import pandas as pd
import os
import glob
from datetime import datetime
import json

logger = logging.getLogger(__name__)

def validate_step31_complete(
    output_base_dir: str = "tests/test_output"
) -> Dict[str, Any]:
    """
    Validates the outputs of Step 31: Gap Analysis Workbook Generator.
    Checks for Excel workbook, CSV data, executive summary, and coverage matrix.
    """
    validation_results = {
        "validation_passed": False,  # Default to False - requires real data validation
        "files": {},
        "errors": ["No real data validation performed"],
        "warnings": ["Using default validation - real data required"],
        "statistics": {
            "validation_type": "default",
            "data_source": "unknown",
            "reliability": "low"
        }
    }

    output_dir = Path(output_base_dir)
    
    # Expected output files (with period patterns)
    expected_patterns = {
        "gap_workbook_excel": f"{output_dir}/gap_analysis_workbook_*.xlsx",
        "gap_workbook_csv": f"{output_dir}/gap_analysis_workbook_data_*.csv",
        "executive_summary": f"{output_dir}/gap_workbook_executive_summary_*.json",
        "cluster_coverage_matrix": f"{output_dir}/cluster_coverage_matrix_*.csv"
    }

    for file_type, pattern in expected_patterns.items():
        candidate_files = sorted(glob.glob(str(pattern)), key=os.path.getctime, reverse=True)
        
        file_validation = {"exists": False, "count": len(candidate_files)}
        
        if candidate_files:
            file_path = Path(candidate_files[0])
            file_validation["exists"] = True
            file_validation["latest"] = str(file_path)
            
            if file_path.stat().st_size == 0:
                validation_results["validation_passed"] = False
                validation_results["warnings"].append(f"Empty file found for {file_type}: {file_path.name}")
                file_validation["size_bytes"] = 0
            else:
                file_validation["size_bytes"] = file_path.stat().st_size
                file_validation["size_mb"] = round(file_path.stat().st_size / (1024 * 1024), 3)
                
                # File-specific validation
                if file_path.suffix == '.csv':
                    try:
                        df = pd.read_csv(file_path)
                        file_validation["rows"] = len(df)
                        file_validation["columns"] = list(df.columns)
                        
                        if df.empty:
                            validation_results["validation_passed"] = False
                            validation_results["warnings"].append(f"CSV file is empty: {file_path.name}")
                        else:
                            # Validate specific columns based on file type
                            if file_type == "gap_workbook_csv":
                                expected_cols = ["cluster_id", "dimension", "coverage_score", "gap_severity"]
                                missing_cols = [col for col in expected_cols if col not in df.columns]
                                if missing_cols:
                                    validation_results["warnings"].append(f"Missing expected columns in {file_type}: {missing_cols}")
                                
                                # Validate coverage data
                                if "coverage_score" in df.columns:
                                    if not df["coverage_score"].dtype in ['int64', 'float64']:
                                        validation_results["warnings"].append("coverage_score should be numeric")
                                    if (df["coverage_score"] < 0).any() or (df["coverage_score"] > 1).any():
                                        validation_results["warnings"].append("coverage_score should be between 0 and 1")
                                
                                if "gap_severity" in df.columns:
                                    valid_severities = ["CRITICAL", "SIGNIFICANT", "MODERATE", "OPTIMAL"]
                                    invalid_severities = df[~df["gap_severity"].isin(valid_severities)]["gap_severity"].unique()
                                    if len(invalid_severities) > 0:
                                        validation_results["warnings"].append(f"Invalid gap_severity values: {invalid_severities}")
                                
                                file_validation["total_entries"] = len(df)
                                file_validation["unique_clusters"] = df["cluster_id"].nunique() if "cluster_id" in df.columns else 0
                                file_validation["unique_dimensions"] = df["dimension"].nunique() if "dimension" in df.columns else 0
                                
                            elif file_type == "cluster_coverage_matrix":
                                expected_cols = ["cluster_id", "category_coverage", "price_band_coverage", "style_orientation", "product_role_balance", "seasonal_capacity", "overall_coverage"]
                                missing_cols = [col for col in expected_cols if col not in df.columns]
                                if missing_cols:
                                    validation_results["warnings"].append(f"Missing expected columns in {file_type}: {missing_cols}")
                                
                                # Validate coverage matrix data
                                coverage_cols = ["category_coverage", "price_band_coverage", "style_orientation", "product_role_balance", "seasonal_capacity", "overall_coverage"]
                                for col in coverage_cols:
                                    if col in df.columns:
                                        if not df[col].dtype in ['int64', 'float64']:
                                            validation_results["warnings"].append(f"{col} should be numeric")
                                        if (df[col] < 0).any() or (df[col] > 1).any():
                                            validation_results["warnings"].append(f"{col} should be between 0 and 1")
                                
                                file_validation["total_clusters"] = len(df)
                                file_validation["coverage_dimensions"] = len([col for col in coverage_cols if col in df.columns])
                                
                    except Exception as e:
                        validation_results["validation_passed"] = False
                        validation_results["errors"].append(f"Error reading CSV {file_path.name}: {e}")
                        
                elif file_path.suffix == '.json':
                    try:
                        with open(file_path, 'r') as f:
                            json_data = json.load(f)
                        
                        file_validation["json_keys"] = list(json_data.keys()) if isinstance(json_data, dict) else "Not a dict"
                        
                        # Validate executive summary structure
                        if file_type == "executive_summary":
                            expected_keys = ["workbook_metadata", "coverage_analysis", "gap_summary", "recommendations"]
                            missing_keys = [key for key in expected_keys if key not in json_data]
                            if missing_keys:
                                validation_results["warnings"].append(f"Missing expected keys in {file_type}: {missing_keys}")
                            
                            # Validate coverage analysis
                            if "coverage_analysis" in json_data:
                                coverage = json_data["coverage_analysis"]
                                if "total_clusters_analyzed" in coverage:
                                    if not isinstance(coverage["total_clusters_analyzed"], int):
                                        validation_results["warnings"].append("total_clusters_analyzed should be integer")
                                if "average_coverage_score" in coverage:
                                    if not isinstance(coverage["average_coverage_score"], (int, float)):
                                        validation_results["warnings"].append("average_coverage_score should be numeric")
                                    if coverage["average_coverage_score"] < 0 or coverage["average_coverage_score"] > 1:
                                        validation_results["warnings"].append("average_coverage_score should be between 0 and 1")
                            
                            # Validate gap summary
                            if "gap_summary" in json_data:
                                gap_summary = json_data["gap_summary"]
                                if "critical_gaps" in gap_summary:
                                    if not isinstance(gap_summary["critical_gaps"], int):
                                        validation_results["warnings"].append("critical_gaps should be integer")
                                if "total_gaps_identified" in gap_summary:
                                    if not isinstance(gap_summary["total_gaps_identified"], int):
                                        validation_results["warnings"].append("total_gaps_identified should be integer")
                                
                    except Exception as e:
                        validation_results["validation_passed"] = False
                        validation_results["errors"].append(f"Error reading JSON {file_path.name}: {e}")
                        
                elif file_path.suffix == '.xlsx':
                    try:
                        # Basic Excel file validation (size and existence)
                        file_validation["excel_valid"] = True
                        
                        # Try to read Excel file to validate structure
                        try:
                            import openpyxl
                            workbook = openpyxl.load_workbook(file_path)
                            file_validation["sheet_names"] = workbook.sheetnames
                            file_validation["total_sheets"] = len(workbook.sheetnames)
                            
                            # Check for expected sheets
                            expected_sheets = ["Executive Summary", "Coverage Matrix", "Store Analysis", "Gap Details"]
                            present_sheets = [sheet for sheet in expected_sheets if sheet in workbook.sheetnames]
                            if present_sheets:
                                file_validation["expected_sheets_found"] = present_sheets
                            else:
                                validation_results["warnings"].append(f"Expected Excel sheets not found: {expected_sheets}")
                                
                        except ImportError:
                            validation_results["warnings"].append("openpyxl not available for Excel validation")
                        except Exception as e:
                            validation_results["warnings"].append(f"Error reading Excel file: {e}")
                            
                    except Exception as e:
                        validation_results["validation_passed"] = False
                        validation_results["errors"].append(f"Error validating Excel file {file_path.name}: {e}")
                        
            validation_results["files"][file_type] = file_validation
        else:
            validation_results["validation_passed"] = False
            validation_results["warnings"].append(f"Missing expected Step 31 output: {pattern}")

    # Calculate statistics
    validation_results["statistics"]["files_found"] = sum(1 for f in validation_results["files"].values() if f["exists"])
    validation_results["statistics"]["total_files_expected"] = len(expected_patterns)
    validation_results["statistics"]["validation_timestamp"] = datetime.now().isoformat()
    
    return validation_results

def validate_executive_summary_json(file_path: str) -> Dict[str, Any]:
    """
    Validates the executive summary JSON file specifically.
    """
    validation_results = {
        "validation_passed": False,  # Requires real data validation
        "errors": [],
        "warnings": [],
        "statistics": {}
    }
    
    try:
        with open(file_path, 'r') as f:
            data = json.load(f)
        
        # Check required structure
        required_keys = ["workbook_metadata", "coverage_analysis", "gap_summary", "recommendations"]
        missing_keys = [key for key in required_keys if key not in data]
        if missing_keys:
            validation_results["validation_passed"] = False
            validation_results["errors"].append(f"Missing required keys: {missing_keys}")
            return validation_results
        
        # Validate workbook metadata
        metadata = data.get("workbook_metadata", {})
        if "generation_timestamp" in metadata:
            if not isinstance(metadata["generation_timestamp"], str):
                validation_results["warnings"].append("generation_timestamp should be string")
        
        if "total_clusters_analyzed" in metadata:
            if not isinstance(metadata["total_clusters_analyzed"], int) or metadata["total_clusters_analyzed"] < 0:
                validation_results["warnings"].append("total_clusters_analyzed should be non-negative integer")
        
        # Validate coverage analysis
        coverage = data.get("coverage_analysis", {})
        if "average_coverage_score" in coverage:
            if not isinstance(coverage["average_coverage_score"], (int, float)):
                validation_results["warnings"].append("average_coverage_score should be numeric")
            if coverage["average_coverage_score"] < 0 or coverage["average_coverage_score"] > 1:
                validation_results["warnings"].append("average_coverage_score should be between 0 and 1")
        
        # Validate gap summary
        gap_summary = data.get("gap_summary", {})
        if "critical_gaps" in gap_summary:
            if not isinstance(gap_summary["critical_gaps"], int) or gap_summary["critical_gaps"] < 0:
                validation_results["warnings"].append("critical_gaps should be non-negative integer")
        
        # Statistics
        validation_results["statistics"]["total_keys"] = len(data)
        validation_results["statistics"]["has_metadata"] = "workbook_metadata" in data
        validation_results["statistics"]["has_coverage"] = "coverage_analysis" in data
        validation_results["statistics"]["has_gap_summary"] = "gap_summary" in data
        validation_results["statistics"]["has_recommendations"] = "recommendations" in data
        
    except Exception as e:
        validation_results["validation_passed"] = False
        validation_results["errors"].append(f"Error reading JSON file: {e}")
    
    return validation_results

def validate_coverage_matrix_csv(file_path: str) -> Dict[str, Any]:
    """
    Validates the cluster coverage matrix CSV file specifically.
    """
    validation_results = {
        "validation_passed": False,  # Requires real data validation
        "errors": [],
        "warnings": [],
        "statistics": {}
    }
    
    try:
        df = pd.read_csv(file_path)
        
        # Check required columns
        required_columns = ["cluster_id", "category_coverage", "price_band_coverage", "style_orientation", "product_role_balance", "seasonal_capacity", "overall_coverage"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            validation_results["validation_passed"] = False
            validation_results["errors"].append(f"Missing required columns: {missing_columns}")
            return validation_results
        
        # Validate cluster IDs
        if not df["cluster_id"].dtype in ['int64']:
            validation_results["warnings"].append("cluster_id should be integer")
        else:
            if (df["cluster_id"] < 0).any() or (df["cluster_id"] > 50).any():
                validation_results["warnings"].append("cluster_id should be between 0 and 50")
        
        # Validate coverage scores
        coverage_columns = ["category_coverage", "price_band_coverage", "style_orientation", "product_role_balance", "seasonal_capacity", "overall_coverage"]
        for col in coverage_columns:
            if col in df.columns:
                if not df[col].dtype in ['int64', 'float64']:
                    validation_results["warnings"].append(f"{col} should be numeric")
                else:
                    if (df[col] < 0).any() or (df[col] > 1).any():
                        validation_results["warnings"].append(f"{col} should be between 0 and 1")
        
        # Statistics
        validation_results["statistics"]["total_clusters"] = len(df)
        validation_results["statistics"]["unique_clusters"] = df["cluster_id"].nunique()
        validation_results["statistics"]["avg_overall_coverage"] = df["overall_coverage"].mean() if "overall_coverage" in df.columns else 0
        validation_results["statistics"]["min_coverage"] = df["overall_coverage"].min() if "overall_coverage" in df.columns else 0
        validation_results["statistics"]["max_coverage"] = df["overall_coverage"].max() if "overall_coverage" in df.columns else 0
        
    except Exception as e:
        validation_results["validation_passed"] = False
        validation_results["errors"].append(f"Error reading CSV file: {e}")
    
    return validation_results

if __name__ == "__main__":
    print("Testing Step 31 Gap Analysis Workbook Validator...")
    results = validate_step31_complete("output")
    print(f"Validation passed: {results['validation_passed']}")
    print(f"Files found: {results['statistics']['files_found']}/{results['statistics']['total_files_expected']}")
    if results['errors']:
        print(f"Errors: {results['errors']}")
    if results['warnings']:
        print(f"Warnings: {results['warnings']}")


