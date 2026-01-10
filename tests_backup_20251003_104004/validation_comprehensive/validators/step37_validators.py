#!/usr/bin/env python3
"""
Step 37 Customer Delivery Formatter Validators

This module contains validators for customer delivery formatter validation from Step 37.

Author: Data Pipeline
Date: 2025-01-16
"""

import logging
from pathlib import Path
from typing import Dict, Any, List
import pandas as pd
import os
import glob
from datetime import datetime
import json

logger = logging.getLogger(__name__)

def validate_step37_complete(
    output_base_dir: str = "tests/test_output"
) -> Dict[str, Any]:
    """
    Validates the outputs of Step 37: Customer Delivery Formatter.
    Checks for customer delivery Excel, CSV, and additional files.
    """
    validation_results = {
        "validation_passed": False,  # Default to False - requires real data validation
        "files": {},
        "errors": ["No real data validation performed"],
        "warnings": ["Using default validation - real data required"],
        "statistics": {
            "validation_type": "default",
            "data_source": "unknown",
            "reliability": "low"
        }
    }

    output_dir = Path(output_base_dir)
    
    # Expected output files (with period patterns)
    expected_patterns = {
        "customer_delivery_excel": f"{output_dir}/customer_delivery_*.xlsx",
        "customer_delivery_csv": f"{output_dir}/customer_delivery_*_store_lines.csv"
    }

    for file_type, pattern in expected_patterns.items():
        candidate_files = sorted(glob.glob(str(pattern)), key=os.path.getctime, reverse=True)
        
        file_validation = {"exists": False, "count": len(candidate_files)}
        
        if candidate_files:
            file_path = Path(candidate_files[0])
            file_validation["exists"] = True
            file_validation["latest"] = str(file_path)
            
            if file_path.stat().st_size == 0:
                validation_results["validation_passed"] = False
                validation_results["warnings"].append(f"Empty file found for {file_type}: {file_path.name}")
                file_validation["size_bytes"] = 0
            else:
                file_validation["size_bytes"] = file_path.stat().st_size
                file_validation["size_mb"] = round(file_path.stat().st_size / (1024 * 1024), 3)
                
                # File-specific validation
                if file_path.suffix == '.csv':
                    try:
                        df = pd.read_csv(file_path)
                        file_validation["rows"] = len(df)
                        file_validation["columns"] = list(df.columns)
                        
                        if df.empty:
                            validation_results["validation_passed"] = False
                            validation_results["warnings"].append(f"CSV file is empty: {file_path.name}")
                        else:
                            # Validate specific columns for customer delivery
                            if file_type == "customer_delivery_csv":
                                expected_cols = ["Store_Code", "Action", "Allocated_ΔQty_Rounded", "Target_SPU_Quantity", "Action_Priority_Rank", "Category", "Subcategory"]
                                missing_cols = [col for col in expected_cols if col not in df.columns]
                                if missing_cols:
                                    validation_results["warnings"].append(f"Missing expected columns in {file_type}: {missing_cols}")
                                
                                # Validate store codes
                                if "Store_Code" in df.columns:
                                    if not df["Store_Code"].dtype in ['int64', 'object']:
                                        validation_results["warnings"].append("Store_Code should be integer or string")
                                    else:
                                        if df["Store_Code"].dtype == 'int64':
                                            if (df["Store_Code"] < 10000).any() or (df["Store_Code"] > 99999).any():
                                                validation_results["warnings"].append("Store_Code should be between 10000 and 99999")
                                
                                # Validate action values
                                if "Action" in df.columns:
                                    valid_actions = ["Add", "Reduce", "No-Change"]
                                    invalid_actions = df[~df["Action"].isin(valid_actions)]["Action"].unique()
                                    if len(invalid_actions) > 0:
                                        validation_results["warnings"].append(f"Invalid Action values: {invalid_actions}")
                                
                                # Validate quantity changes
                                if "Allocated_ΔQty_Rounded" in df.columns:
                                    if not df["Allocated_ΔQty_Rounded"].dtype in ['int64', 'float64']:
                                        validation_results["warnings"].append("Allocated_ΔQty_Rounded should be numeric")
                                
                                # Validate target quantities
                                if "Target_SPU_Quantity" in df.columns:
                                    if not df["Target_SPU_Quantity"].dtype in ['int64', 'float64']:
                                        validation_results["warnings"].append("Target_SPU_Quantity should be numeric")
                                    else:
                                        if (df["Target_SPU_Quantity"] < 0).any():
                                            validation_results["warnings"].append("Found negative Target_SPU_Quantity values")
                                
                                # Validate priority rank
                                if "Action_Priority_Rank" in df.columns:
                                    if not df["Action_Priority_Rank"].dtype in ['int64', 'float64']:
                                        validation_results["warnings"].append("Action_Priority_Rank should be numeric")
                                    else:
                                        if (df["Action_Priority_Rank"] < 1).any():
                                            validation_results["warnings"].append("Action_Priority_Rank should be >= 1")
                                
                                file_validation["total_actions"] = len(df)
                                file_validation["unique_stores"] = df["Store_Code"].nunique() if "Store_Code" in df.columns else 0
                                
                                # Analyze action distribution
                                if "Action" in df.columns:
                                    action_dist = df["Action"].value_counts().to_dict()
                                    file_validation["action_distribution"] = action_dist
                                
                                # Analyze category distribution
                                if "Category" in df.columns:
                                    category_dist = df["Category"].value_counts().to_dict()
                                    file_validation["category_distribution"] = category_dist
                                
                    except Exception as e:
                        validation_results["validation_passed"] = False
                        validation_results["errors"].append(f"Error reading CSV {file_path.name}: {e}")
                        
                elif file_path.suffix == '.xlsx':
                    try:
                        # Basic Excel file validation (size and existence)
                        file_validation["excel_valid"] = True
                        
                        # Try to read Excel file to validate structure
                        try:
                            import openpyxl
                            workbook = openpyxl.load_workbook(file_path)
                            file_validation["sheet_names"] = workbook.sheetnames
                            file_validation["total_sheets"] = len(workbook.sheetnames)
                            
                            # Check for expected sheets
                            expected_sheets = ["Overview", "Store Lines", "Cluster Summary", "Top Adds by Cluster", "Women Casual Pants"]
                            present_sheets = [sheet for sheet in expected_sheets if sheet in workbook.sheetnames]
                            if present_sheets:
                                file_validation["expected_sheets_found"] = present_sheets
                            else:
                                validation_results["warnings"].append(f"Expected Excel sheets not found: {expected_sheets}")
                                
                        except ImportError:
                            validation_results["warnings"].append("openpyxl not available for Excel validation")
                        except Exception as e:
                            validation_results["warnings"].append(f"Error reading Excel file: {e}")
                            
                    except Exception as e:
                        validation_results["validation_passed"] = False
                        validation_results["errors"].append(f"Error validating Excel file {file_path.name}: {e}")
                        
            validation_results["files"][file_type] = file_validation
        else:
            validation_results["validation_passed"] = False
            validation_results["warnings"].append(f"Missing expected Step 37 output: {pattern}")

    # Calculate statistics
    validation_results["statistics"]["files_found"] = sum(1 for f in validation_results["files"].values() if f["exists"])
    validation_results["statistics"]["total_files_expected"] = len(expected_patterns)
    validation_results["statistics"]["validation_timestamp"] = datetime.now().isoformat()
    
    return validation_results

def validate_customer_delivery_csv(file_path: str) -> Dict[str, Any]:
    """
    Validates the customer delivery CSV file specifically.
    """
    validation_results = {
        "validation_passed": False,  # Requires real data validation
        "errors": [],
        "warnings": [],
        "statistics": {}
    }
    
    try:
        df = pd.read_csv(file_path)
        
        # Check required columns
        required_columns = ["Store_Code", "Action", "Allocated_ΔQty_Rounded", "Target_SPU_Quantity", "Action_Priority_Rank", "Category", "Subcategory"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            validation_results["validation_passed"] = False
            validation_results["errors"].append(f"Missing required columns: {missing_columns}")
            return validation_results
        
        # Validate store codes
        if not df["Store_Code"].dtype in ['int64', 'object']:
            validation_results["warnings"].append("Store_Code should be integer or string")
        else:
            if df["Store_Code"].dtype == 'int64':
                if (df["Store_Code"] < 10000).any() or (df["Store_Code"] > 99999).any():
                    validation_results["warnings"].append("Store_Code should be between 10000 and 99999")
        
        # Validate action values
        valid_actions = ["Add", "Reduce", "No-Change"]
        invalid_actions = df[~df["Action"].isin(valid_actions)]["Action"].unique()
        if len(invalid_actions) > 0:
            validation_results["warnings"].append(f"Invalid Action values: {invalid_actions}")
        
        # Validate quantity changes
        if not df["Allocated_ΔQty_Rounded"].dtype in ['int64', 'float64']:
            validation_results["warnings"].append("Allocated_ΔQty_Rounded should be numeric")
        
        # Validate target quantities
        if not df["Target_SPU_Quantity"].dtype in ['int64', 'float64']:
            validation_results["warnings"].append("Target_SPU_Quantity should be numeric")
        else:
            if (df["Target_SPU_Quantity"] < 0).any():
                validation_results["warnings"].append("Found negative Target_SPU_Quantity values")
        
        # Validate priority rank
        if not df["Action_Priority_Rank"].dtype in ['int64', 'float64']:
            validation_results["warnings"].append("Action_Priority_Rank should be numeric")
        else:
            if (df["Action_Priority_Rank"] < 1).any():
                validation_results["warnings"].append("Action_Priority_Rank should be >= 1")
        
        # Check for missing critical data
        if df["Store_Code"].isnull().any():
            validation_results["warnings"].append("Found missing Store_Code values")
        
        if df["Action"].isnull().any():
            validation_results["warnings"].append("Found missing Action values")
        
        if df["Category"].isnull().any():
            validation_results["warnings"].append("Found missing Category values")
        
        # Statistics
        validation_results["statistics"]["total_actions"] = len(df)
        validation_results["statistics"]["unique_stores"] = df["Store_Code"].nunique()
        validation_results["statistics"]["action_distribution"] = df["Action"].value_counts().to_dict()
        validation_results["statistics"]["category_distribution"] = df["Category"].value_counts().to_dict()
        validation_results["statistics"]["avg_target_quantity"] = df["Target_SPU_Quantity"].mean() if "Target_SPU_Quantity" in df.columns else 0
        validation_results["statistics"]["total_quantity_change"] = df["Allocated_ΔQty_Rounded"].sum() if "Allocated_ΔQty_Rounded" in df.columns else 0
        
    except Exception as e:
        validation_results["validation_passed"] = False
        validation_results["errors"].append(f"Error reading CSV file: {e}")
    
    return validation_results

if __name__ == "__main__":
    print("Testing Step 37 Customer Delivery Formatter Validator...")
    results = validate_step37_complete("output")
    print(f"Validation passed: {results['validation_passed']}")
    print(f"Files found: {results['statistics']['files_found']}/{results['statistics']['total_files_expected']}")
    if results['errors']:
        print(f"Errors: {results['errors']}")
    if results['warnings']:
        print(f"Warnings: {results['warnings']}")


